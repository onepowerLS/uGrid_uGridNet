# -*- coding: utf-8 -*-
"""
uGrid "Macro" Code

Description: This runs the uGrid tool to provide optimization of equipment sizes by minimizing
the levelized cost of electricity, also referred to as the tariff. The optimization is a
particle swarm optimization adjusted for constraints specific to this optimization problem. 

Input: Four spreadsheets need to be in the file folder: main inputs (uGrid_Input.xlsx), 
load data (load.xlsx), load forecasting data (fullyearenergy.xlsx), and 
weather data (TMY.xlsx). See the uGrid User Guide for more information.

Output: Results from each generation will be output to the command line. The 
generation results and overall results will be output to an excel spreadsheet 
with the name specified in the main inputs spreadsheet.

How to Run: Update the data in the four input spreadsheets. Run this file.  

@author: Phylicia Cicilio
"""

from __future__ import division
import glob
import numpy as np
import pandas as pd
from technical_tools_PC_3_alt import Tech_total
from economic_tools_PC_3 import Econ_total
import time
import datetime as dt
from constants import SITE_NAME
from openpyxl import Workbook, load_workbook

def get_8760(village_name):
    filtered_list = glob.glob(f'{village_name}*8760*.xlsx')
    for f in filtered_list:
        if village_name in f and '8760' in f:
            return f
    return None

if __name__ == "__main__":
    close_fds=False #subprocess management

    #Recording time to complete
    start_time = time.time()
    sitename =SITE_NAME
    #Call in Inputs
    #PSO Parameters
    PSO_Parameters = pd.read_excel(sitename +'_uGrid_Input.xlsx', sheet_name = 'PSO')
    #TODO: Here
    maxGen = PSO_Parameters['maxGen'][0]
    numInd = PSO_Parameters['numInd'][0]
    X_tariff_multiplier = PSO_Parameters['X_tariff_multiplier'][0]
    stopLimit = PSO_Parameters['stopLimit'][0]
    convergenceRequirement = PSO_Parameters['convergenceRequirement'][0]
    lowTestLim = PSO_Parameters['lowTestLim'][0]
    highTestLim = PSO_Parameters['highTestLim'][0]
    roundDownSize = PSO_Parameters['roundDownSize'][0]
    C1 = PSO_Parameters['C1'][0]
    C2 = PSO_Parameters['C2'][0]
    CF = PSO_Parameters['CF'][0]
    W = PSO_Parameters['W'][0]
    VF = PSO_Parameters['VF'][0]

    #Parameter limits: Battery, PV
    #These will be changed so the solutions are more flexible for sites (not input)
    #Could make these scaled off of load input
    lower_bounds = [1,1]
    upper_bounds = [25,5]
    
    
    #Initialize matrixes and parameters
    Parameters_test = np.zeros(2)
    Parameters_dev = np.zeros(2)   
    Parameters = np.zeros((2,numInd,maxGen)) #this is the result for each parameter, for each month, for each individual
    Propane_ec = np.zeros((numInd,maxGen))
    Batt_kWh_tot_ec = np.zeros((numInd,maxGen))

    # Capex variables
    bank_opt = [] 
    panels_opt = []
    tracker_opt = []
    BOS_opt = []
    
    #Create random initial guesses for Parameters
    for i in range(2):
        for k in range(numInd):
            rn = np.random.uniform(lower_bounds[i],upper_bounds[i])
            if rn < roundDownSize: #Constraint for minimum sizes
                rn = 0
                Parameters[i,k,0] = np.copy(rn)
            else:
                Parameters[i,k,0] = np.copy(rn)
    
    #Initialize Economic Parameters    
    tariff = np.zeros((numInd,maxGen))
    Batt_life_yrs = np.zeros((numInd,maxGen))
    
    #Initialize Global Bests
    #global best: best known postions ever, personal best: best known position for each particle(out of all generations), gbest_change: best in generation (out of all individuals)
    #Global Best
    gB_propane = 999
    gB_tariff = 999
    gB_tariff_plus = gB_tariff*(1+X_tariff_multiplier)
    gB_parameters = np.zeros(2)
    loadfile = get_8760(sitename)
    load = pd.read_excel(loadfile, sheet_name='8760', usecols='B')
    #TODO: Here
    hmax = len(load)
    gB_Cost = np.zeros(hmax)
    data_plot_variables = np.zeros((hmax,6))
    gB_plot_variables = pd.DataFrame(data = data_plot_variables,columns=['Batt_SOC', 'LoadkW', 'P_gen', 'P_PV', 'P_batt', 'P_dump'])
    #print(gB_plot_variables)
    data_optimization_variables = np.zeros((maxGen-1, 4))
    gB_optimization_output_var = np.zeros((4,maxGen-1))
    gB_optimization_output =pd.DataFrame(data = data_optimization_variables, columns =['BattkW','PVkW','Propane','Tariff'])
    #print(gB_optimization_output)           

    #gB_change (best individual in generation)
    gB_change_propane = np.ones(maxGen)*999
    gB_change_tariff = np.ones(maxGen)*999
    gB_change_tariff_plus = gB_change_tariff*(1+X_tariff_multiplier)
    gB_change_parameters = np.zeros((2,maxGen))
    #print(gB_change_parameters)
    #TODO: add gB_change_cost
    gB_change_cost = np.ones(maxGen)
    gB_change_capex = np.ones((7,maxGen))
    #print(gB_tariff, gB_Cost,gB_change_tariff, gB_change_cost)
    #Personal Best (best position for each particle)
    pB_propane = np.ones(numInd)*999
    pB_tariff = np.ones(numInd)*999
    pB_parameters = np.ones((2,numInd))*999
    pB_tariff_plus = np.ones(numInd)*999

    #Initialize Velocities
    #batt, PV, CSP, ORC, TES
    VelMax = np.zeros(2)
    VelMin = np.zeros(2)
    VelMax = VF*(np.array(upper_bounds)-np.array(lower_bounds))
    VelMin = -1*VelMax
    V_gens = np.zeros((2,numInd,maxGen))

    #Record History (best tariff recording through iterations)
    recordRecord = np.zeros(maxGen)
    

    # Start Optimization Iterations
    iteration = 0
    t_diff = 999
    match = 0
    while iteration < maxGen-1 and t_diff > stopLimit and match < convergenceRequirement:
        match = 0
        for m in range(numInd):            
            #calculate technical parameters
            Propane_ec[m,iteration], Batt_SOC, LoadkW, P_gen, P_PV, P_batt, P_dump,Limit_charge, Limit_discharge, BattkW, Batt_kWh_tot_ec[m,iteration],loadkWh,peakload = Tech_total(Parameters[0,m,iteration],Parameters[1,m,iteration])
            print('Battery KW: ' + str(BattkW))
            #don't need to save Gt_panel, final, Batt_SOC, and Charge these are used to validate program
            LoanPrincipal, year, Cost, Revenue, CashonHand, Balance, M, O, tariff[m,iteration], Batt_life_yrs[m,iteration], Cost_EPC, Cost_Dist, Cost_BOS, Cost_bank, C1_LPG, Cost_inv, Cost_panels, Cost_EPC_tracker, Cost_labour, C1_pv, PVkW= Econ_total(Propane_ec[m,iteration],Parameters[1,m,iteration]*peakload,Parameters[0,m,iteration]*peakload,Batt_kWh_tot_ec[m,iteration],peakload,loadkWh)
            #order of parameters: batt, PV, CSP, ORC, TES_ratio
        
   
            #Checking Outputs
            print("This individual's yearly propane and tariff is: " +str(Propane_ec[m,iteration]) + ", " +str(tariff[m,iteration]))
            #print("peakload = " +str(peakload))
            print("cost = "+str(sum(Cost)))
            #C1_pv = Cost_panels + Cost_BOS + Cost_EPC + Cost_Dev
            print('Cost EPc '+str(Cost_EPC),'Cost Distribution'+ str(Cost_Dist),'Cost BOS' +str(Cost_BOS),'Cost bank' +str(Cost_bank),'Cost Genset' + str(C1_LPG),'Cost inverter'+ str(Cost_inv),'Cost panels'+ str(Cost_panels),'Cost tracker'+ str(Cost_EPC_tracker), 'PVkW ;' + str(PVkW))
            print('C1_pv is :' + str(C1_pv))
            
            #Find generation best
            if tariff[m,iteration] < gB_change_tariff_plus[iteration] or (tariff[m,iteration] < gB_change_tariff[iteration] and Propane_ec[m,0] < gB_change_propane[iteration]):
                gB_change_tariff[iteration] = np.copy(tariff[m,iteration])
                gB_change_tariff_plus[iteration] = gB_change_tariff[iteration]*(1+X_tariff_multiplier)
                gB_change_propane[iteration] = np.copy(Propane_ec[m,iteration])
                gB_change_parameters[:,iteration] = np.copy(Parameters[:,m,iteration])
                gB_change_cost[iteration] = np.copy(Cost[iteration])
                
                #print(gB_change_cost[iteration], )
            #Find global best (removed gB_tariff_plus buffer for true global best: tariff[m,iteration] < gB_tariff_plus or)
            if tariff[m,iteration] < gB_tariff_plus or (tariff[m,iteration] < gB_tariff and Propane_ec[m,0] < gB_propane):
                gB_tariff = np.copy(tariff[m,iteration])
                gB_tariff_plus = gB_tariff*(1+X_tariff_multiplier)
                gB_propane = np.copy(Propane_ec[m,iteration])
                gB_parameters = np.copy(Parameters[:,m,iteration])
                #Saving plotting variables
                data_plot_variables = np.column_stack((Batt_SOC[0:8760], LoadkW, P_gen, P_PV, P_batt, P_dump))
                #print(data_plot_variables)
                #data_plot_variables =[Batt_SOC, LoadkW, P_gen, P_PV, P_batt, P_dump]
                gB_plot_variables = pd.DataFrame(data_plot_variables,columns=['Batt_SOC', 'LoadkW', 'P_gen', 'P_PV', 'P_batt', 'P_dump'])
                gB_Cost = np.copy(Cost)
            #Find personal best (done at the end of each iteration)
            for pB_iter in range(iteration):
                if tariff[m,iteration] < pB_tariff_plus[m] or (tariff[m,iteration] < pB_tariff[m] and Propane_ec[m,iteration] < pB_propane[m]):
                    pB_tariff[m] = np.copy(tariff[m,iteration])
                    pB_tariff_plus[m] = pB_tariff[m]*(1+X_tariff_multiplier)
                    pB_propane[m] = np.copy(Propane_ec[m,iteration])
                    pB_parameters[:,m] = np.copy(Parameters[:,m,iteration])
            recordRecord[iteration] = np.copy(gB_tariff)

            #Calculate Next Gen
            V_gens[:,m,iteration+1] = W*V_gens[:,m,iteration] + C1*np.random.uniform(0,1)*(pB_parameters[:,m] - Parameters[:,m,iteration]) + C2*np.random.uniform(0,1)*(gB_parameters- Parameters[:,m,iteration])
            #Check that velocity is within limits
            for s in range(2):
                if V_gens[s,m,iteration+1] > VelMax[s]:
                    V_gens[s,m,iteration+1] = np.copy(VelMax[s])
                if V_gens[s,m,iteration+1] < VelMin[s]:
                    V_gens[s,m,iteration+1] = np.copy(VelMin[s])
            # new microgrid parameters for the next generation
            # value from previous position plus some movement based on velocity & PSO parameters
            Parameters[:,m,iteration+1] = Parameters[:,m,iteration] + CF*V_gens[:,m,iteration+1]
            # ceilings based on limits for specific microgrid parameters
            for q in range(2):
                if Parameters[q,m,iteration+1] > upper_bounds[q]:
                    Parameters[q,m,iteration+1] = np.copy(upper_bounds[q])
                if Parameters[q,m,iteration+1] < lower_bounds[q]:
                    Parameters[q,m,iteration+1] = np.copy(lower_bounds[q])
                if Parameters[q,m,iteration+1] < roundDownSize:
                    Parameters[q,m,iteration+1] = 0
           
        #Stopping Criteria (once all individuals are calculated)
        #Testing is the equipment size parameters match
        if iteration < 30:
            testLim = np.copy(lowTestLim)
        else:
            testLim = np.copy(highTestLim) #more restrictive stopping criteria can be placed for later generations
        match = 0
        for ind in range(numInd):
            Parameters_test = (Parameters[:,ind,iteration] - gB_parameters)/(gB_parameters + 0.0001)
            Parameters_dev = np.zeros(2)
            for r in range(2):
                if abs(Parameters_test[r]) <= testLim:
                    Parameters_dev[r] = 1
            PSO_match = sum(Parameters_dev)
            if PSO_match == 2:
                match = match + 1 #equivalent of previous code
        if match > convergenceRequirement:
            print("Stopping due to matching parameter values")
    
        #Checking for minimal change between bests
        if iteration >= 2:
            diff1 = abs(gB_change_tariff[iteration-2]-gB_change_tariff[iteration-1])
            diff2 = abs(gB_change_tariff[iteration-1] - gB_change_tariff[iteration])
            t_diff = diff1+diff2
            if t_diff < stopLimit:
                print("Stopping due to minimal change between global bests")
    
        #Print generation results
        print("Global Best Tariff "+ str(gB_tariff))
        print("Best Tariff in Generation " + str(gB_change_tariff[iteration]))
        print("Propane meeting best tarrif " + str(gB_change_propane[iteration]) + str(gB_propane))
        print("Total Cost of Generation " + str(sum(gB_Cost)))
        iteration += 1

    #Calculate total run time
    end_time = time.time()
    total_time = end_time - start_time
    print("Time to complete simulation is " + str(total_time))

    #Best Solution Variables Saved
    Total_Cost = sum(gB_Cost)
    #add extra variables to solution output
    #This raises the error: ValueError: If using all scalar values, you must pass an index
    gB_total_var =pd.DataFrame({'Total Cost':[Total_Cost],'Simulation_Time':[total_time], 'Peak Load':[peakload]})
    print(gB_total_var)

    #gB_optimization_output_var = np.zeros((4,maxGen-1))
    
    #optimum capex breakdown
    Size_costing_parameters = pd.read_excel(sitename + '_uGrid_Input.xlsx', sheet_name = 'Sizing_Costing', usecols='F')
    Cost_panels_opt_array = gB_change_parameters[1,:]*peakload*Size_costing_parameters.iloc[0].to_numpy()
    Cost_bank_opt_array = gB_change_parameters[0,:]*peakload*Size_costing_parameters.iloc[1].to_numpy()
    Cost_inv_opt_array = peakload*Size_costing_parameters.iloc[2].to_numpy()
    Cost_EPC_tracker_opt_array = gB_change_parameters[1,:]*peakload*Size_costing_parameters.iloc[3].to_numpy()
    C1_LPG_opt_array = peakload*Size_costing_parameters.iloc[4].to_numpy()
    Cost_BOS_opt_array = gB_change_parameters[1,:]*peakload*Size_costing_parameters.iloc[5].to_numpy()
    print(Cost_panels_opt_array,Cost_bank_opt_array, Cost_inv_opt_array, Cost_EPC_tracker_opt_array, C1_LPG_opt_array,Cost_BOS_opt_array)

    for n in range(len(Cost_panels_opt_array)):
       

        if Cost_panels_opt_array[n]>0.0:
            #print("Optimal cost panels" +str(Cost_panels_opt_array[n]),"Optimal cost Panels" +str(Cost_bank_opt_array[n]), "Optimal cost Inverter" +str(Cost_inv_opt_array[n]), "Optimal cost Tracker" +str(Cost_EPC_tracker_opt_array[n]), "Optimal cost GENSET" +str(C1_LPG_opt_array[n]),"Optimal cost BOS" +str(Cost_BOS_opt_array[n]))
            panels_opt.append(Cost_panels_opt_array[n])
            bank_opt.append(Cost_bank_opt_array[n])
            tracker_opt.append(Cost_EPC_tracker_opt_array[n])
            BOS_opt.append(Cost_BOS_opt_array[n])
    print("Optimal cost panels" +str(panels_opt[-1]),"Optimal cost Panels" +str(bank_opt[-1]), "Optimal cost Inverter" +str(Cost_inv_opt_array[-1]), "Optimal cost Tracker" +str(tracker_opt[-1]), "Optimal cost GENSET" +str(C1_LPG_opt_array[-1]),"Optimal cost BOS" +str(BOS_opt[-1]))
    #Cost_bank[iteration],
    #C1_LPG[iteration],
    #Cost_inv[iteration],
    #Cost_EPC_opt[iteration],
    #Cost_EPC_tracker_opt 
    gB_optimization_output_var = {'BattkW':gB_change_parameters[0,:],'PVkW':gB_change_parameters[1,:], 'Propane':gB_change_propane,'Tariff':gB_change_tariff, 'Cost':sum(gB_Cost)}
    #gB_optimization_costs_breakdown = {'Propane':1.3*gB_change_propane}
    #gB_optimization_costs_breakdown = {'Propane':gB_change_propane, 'PV':gB_propane, 'Battery':, 'Inverter':, 'Tracker':, 'Genset':, 'Reticulation':,'EPC':, 'BOS':, 'Labour':, }
    gB_recordRecord = pd.DataFrame({'recordRecord':recordRecord})
    #gB_optimization_output_var = np.transpose(np.concatenate((gB_change_parameters[0,:],gB_change_parameters[1,:], gB_change_propane,gB_change_tariff), axis=1))
    gB_optimization_output =pd.DataFrame(gB_optimization_output_var)#, columns =['BattkW','PVkW','Propane','Tariff'])            
    #print(gB_optimization_output)
    #gB_optimization_costs = pd.DataFrame(gB_optimization_costs_breakdown)

    #Print Results to Excel (Optimization and Solution Variables)
    gB_Size_costing_parameters = pd.read_excel(sitename + '_uGrid_Input.xlsx', sheet_name = 'Sizing_Costing')
    #del gB_Size_costing_parameters['Subtotal', 'QTY']
    #print(gB_Size_costing_parameters)

    filename_xlsx = sitename + "_uGrid_Output_"+PSO_Parameters['output_name'][0]+".xlsx"
    #list = [Cost_EPC, Cost_Dist, Cost_BOS, Cost_bank, C1_LPG, Cost_inv, Cost_panels, Cost_EPC_tracker, C1_pv]
    #gB_cost_parameters = pd.DataFrame(list, columns = ['Subtotal'])
    #print(gB_cost_parameters)
    #gB_Size_costing_parameters.append(gB_cost_parameters)
    
    #TODO: Here
    writer = pd.ExcelWriter(filename_xlsx, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    gB_plot_variables.to_excel(writer, sheet_name='Solution Output')
    gB_optimization_output.to_excel(writer, sheet_name='Optimization Output')
    gB_total_var.to_excel(writer, sheet_name='Other Solution Output')
    gB_recordRecord.to_excel(writer, sheet_name='Record History')
    #gB_optimization_costs.to_excel(writer,sheet_name='Costs_BreakDown')
    gB_Size_costing_parameters.to_excel(writer,sheet_name='Sizing_Costing')
    




    # Close the Pandas Excel writer and output the Excel file.
    writer.save()

    #Plot power to the load and the load curve
    #gB_plot_variables.plot(y = ['LoadkW', 'genLoad', 'Batt_Power_to_Load', 'PV_Power','dumpload'], kind='line')
    #filename_Plot_LoadDispatch = "Power_to_Load_"+PSO_Parameters['output_name'][0]+".pdf"
    #plt.savefig(filename_Plot_LoadDispatch)
    #Plot power to battery and battery SOC
    #gB_plot_variables.plot(y = ['Batt_SOC', 'Charge', 'PV_Batt_Change_Power','Batt_Power_to_Load_neg', 'Batt_frac', 'Gen_Batt_Charge_Power'], kind='line')
    #filename_Plot_BatteryDispatch = "Battery_to_Load_"+PSO_Parameters['output_name'][0]+".pdf"
    #plt.savefig(filename_Plot_BatteryDispatch)


    #write these values in the subtotals column, H
    simdate = dt.datetime.today()
    add0 = lambda x: '0'+str(x) if x < 10 else str(x)
    wb = load_workbook('RAL_uGrid_Output_alt.xlsx')
    ws = wb["Sizing_Costing"]
    
    ws['H2'] = panels_opt[-1]
    ws['H3'] = bank_opt[-1]
    ws['H4'] = Cost_inv[-1]
    ws['H5'] = tracker_opt[-1]
    ws['H6'] = C1_LPG[-1]
    ws['H7'] = BOS_opt[-1]
    ws['H8'] = Cost_Dist[0]
    ws['H9'] = Cost_EPC[0]*10 
    ws['H10'] = Cost_EPC[0]
    ws['H11']= Cost_labour
    ws['H12'] = C1_pv[0]
    ws['H13'] = Total_Cost

    gB_change_parameters_panels = []
    gB_change_parameters_bank =[]

    for i in range(len(gB_change_parameters[1,:])):
        if gB_change_parameters[1][i]>0.0:
            gB_change_parameters_panels.append(gB_change_parameters[1][i])
            gB_change_parameters_bank.append(gB_change_parameters[0][i])
    
    print(gB_change_parameters_panels,gB_change_parameters_bank)

    ws['E2'] = gB_change_parameters_panels[-1]*peakload
    ws['E3'] = gB_change_parameters_bank[-1]*peakload
    ws['E4'] = peakload
    ws['E5'] = gB_change_parameters_panels[-1]*peakload
    ws['E6'] = peakload
    ws['E7'] = gB_change_parameters_panels[-1]*peakload
    ws['E8'] = ' '
    ws['E10'] = ' '
    ws['E11']= ' '
    ws['E12'] = ' ' 
    ws['E13'] = ' '

    wb
    
    wb.save('RAL_uGrid_Output_' + str(simdate.year) + add0(simdate.month) + add0(simdate.day) + \
        '_' + add0(simdate.hour) + add0(simdate.minute) + '.xlsx')
    print("File appended.")

  

    



