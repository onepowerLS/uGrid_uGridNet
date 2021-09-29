"""
Created on Thu Jan 10 09:20:26 2019

uGrid Net
The goal is to take the kml file which has highlighted polygons of "can't build here"
areas and also an inputted generation station gps location, and determine where
to place distribution poles and how to connect the distribution network together.

This creates a LV 220V and MV 6.3kV network layout, where the MV works as a backbone.

Including N-1 Reliability Cost-Benefit of Wiring Layout

@author: Phy

"""

import os
import numpy as np
import pandas as pd
import geopandas as gpd
import math as m
import networkx as nx
import datetime as dt
from pdf2image import convert_from_path
from PIL import Image
from shapely.geometry import Point, LineString
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as img
import time
import matplotlib.pyplot as plt


#==============================================================================
#Calculate Distance between GPS coordinates with Haversine Formula (returns in m)
def GPStoDistance(Lat1,Lat2,Long1,Long2):
    R_earth = 6371 #earth's mean radius in km
    a = m.sin((Lat1-Lat2)/2)**2 + m.cos(Lat1)*m.cos(Lat2)*m.sin((Long1-Long2)/2)**2
    c = 2*m.atan2(m.sqrt(a),m.sqrt(1-a))
    d = (R_earth*c)*1000 #m
    return d
#===============================================================================


#==============================================================================
# Exclusion Mapper using 2nd array instead of increasing list of indexes
def ExclusionMapper(ExclusionMap_array,reformatScaler,exclusionBuffer,d_EW_between,d_NS_between,width_new,height_new):
# This function takes the exclusion array, collected from the pdf image, and 
# recasts the exclusion zones into a new array accounting the for the exclusion 
# buffer and reformat scaler. The reformat scaler is used to reduce the 
# dimensions of the exclusion array so the program can run faster. 
# Additionally, it can make sense to reduce the resolution if the original pixel 
# resolution is less than a foot, such a high resolution is unnecessary. 
    print("in exclusion mapper")
    #Extend the exclusions to include the buffer zone
    bufferArrayWidth_EW = m.ceil(exclusionBuffer/d_EW_between) #reformatted indexes array width 
    bufferArrayHeight_NS = m.ceil(exclusionBuffer/d_NS_between)
    height_og = len(ExclusionMap_array[:,0,0])
    width_og = len(ExclusionMap_array[0,:,0])
    new_exclusions = np.zeros((width_new,height_new))
    indexes = []
    #t0 = time.time()
    
    for i in range((width_og-1),-1,-1):
        k = int((width_og-1-i)/reformatScaler)
        #print(str(i))
        #t1 = time.time()
        #print(str(t1-t0))
        for j in range(0,height_og,1):
            l = int(j/reformatScaler)
            #print(str(i)+" "+str(j))
            if ExclusionMap_array[j,i,0] != 255: #exclusion zones are both grey and black, safe zones are white (255,255,255)
                #set everything within buffer to 1
                if k < bufferArrayWidth_EW:
                    if l < bufferArrayHeight_NS:
                        new_exclusions[0:k+bufferArrayWidth_EW,0:l+bufferArrayHeight_NS] = 1
                    elif l > height_og-bufferArrayHeight_NS:
                        new_exclusions[0:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:height_new] = 1
                    else:
                        new_exclusions[0:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:l+bufferArrayHeight_NS] = 1
                elif k > width_new-bufferArrayWidth_EW:
                    if l < bufferArrayHeight_NS:
                        new_exclusions[k-bufferArrayWidth_EW:width_new,0:l+bufferArrayHeight_NS] = 1
                    elif l > height_og-bufferArrayHeight_NS: 
                        new_exclusions[k-bufferArrayWidth_EW:width_new,l-bufferArrayHeight_NS:height_new] = 1
                    else:
                        new_exclusions[k-bufferArrayWidth_EW:width_new,l-bufferArrayHeight_NS:l+bufferArrayHeight_NS] = 1
                else:
                    if l < bufferArrayHeight_NS:
                        new_exclusions[k-bufferArrayWidth_EW:k+bufferArrayWidth_EW,0:l+bufferArrayHeight_NS] = 1
                    elif l > height_og-bufferArrayHeight_NS: 
                        new_exclusions[k-bufferArrayWidth_EW:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:height_new] = 1
                    else:
                        new_exclusions[k-bufferArrayWidth_EW:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:l+bufferArrayHeight_NS] = 1
    del ExclusionMap_array
    print("finished remapping")
    new_exclusions = np.flip(new_exclusions)
    
    for o in range(width_new):
        for p in range(height_new):
            if new_exclusions[o,p] == 1:
                indexes.append([o,p])
    indexes = np.array(indexes)
    print("finished new indexing")

    #Save new indexes
    index_csv_name = "indexes_reformatted_%s_bufferzone_%s.csv" %(str(reformatScaler),str(exclusionBuffer))
    np.savetxt(index_csv_name,indexes, delimiter=",")
    return indexes
#==============================================================================



#=============================================================================
# Get Distance between array indexes
def DistanceBWindexes(indexesA, indexesB, d_EW_between, d_NS_between):
# This function calculates the distances between array indexes. This works for 
# any array, just the distance between indexes (or “pixels”) needs to be inputted.
    if len(indexesA) == 2:
        #This is a single index submission
        A_sqr = m.pow(((indexesA[0]-indexesB[0])*d_EW_between),2)
        B_sqr = m.pow(((indexesA[1]-indexesB[1])*d_NS_between),2)
    else:
        #This is multiple indexes
        Ax, Bx = np.meshgrid(indexesA[:, 0], indexesB[:, 0])
        Ay, By = np.meshgrid(indexesA[:, 1], indexesB[:, 1])
        
        A_sqr = np.power((Ax.T - Bx.T)*d_EW_between, 2)
        B_sqr = np.power((Ay.T - By.T)*d_NS_between, 2)
    DistanceAB = np.sqrt(A_sqr+B_sqr)
    return DistanceAB
#=============================================================================

    
        
#==============================================================================
# Clustering of Connections for Initial Solution Pole Placement
def Clustering(indexes_conn, num_clusters):
# This function clusters the house location to obtain initial placements of 
# the LV (220V) poles. The number of clusters to create is set by the 
# num_clusters input. The clustering is performed with gaussian mean clustering 
# using the scikitlearn GaussianMixture function. 
    from sklearn import mixture
    
    X = np.copy(indexes_conn)

    cv_type = 'tied'
    gmm = mixture.GaussianMixture(n_components = num_clusters,covariance_type=cv_type)
    gmm.fit(X)
    clf = gmm 
    Y_ = clf.predict(X) #Y_ the index is the connection # and the value is the pole
    #gmm.means_: the index is the pole #, the numbers are the indexes of that pole
    
    #round and convert means to integers
    means = np.copy(gmm.means_) #the location (indexes) of the mean center of the cluster, where the pole is placed
    for i in range(len(gmm.means_)):
        means[i][0] = int(means[i][0])
        means[i][1] = int(means[i][1])
    
    return Y_ ,means
#==============================================================================


#==============================================================================
# Find non-exlcusion spot in growing circular manner
def FindNonExclusionSpot(index_x_og, index_y_og, index_excl_comp,range_limit,max_y,max_x):
# This function determines if the initial pole placement is in an exclusion zone. 
# If the pole is in an exclusion zone, the function tests the area around the 
# initial placement in a circular fashion to find the closest index location 
# to the initial placement that is not an exclusion zone.     
    for i in range(range_limit):
        #Define limits
        x_min = int(index_x_og - i)
        if x_min < 0:
            x_min = 0
        x_max = int(index_x_og + i)
        if x_max > max_x:
            x_max = max_x
        y_min = int(index_y_og - i)
        if y_min < 0:
            y_min = 0
        y_max = int(index_y_og + i)
        if y_max > max_y:
            y_max = 0
        #Testing around spot
        #xmin and xmax
        for y in range(y_min,y_max):
            index_pole_comp = x_min + (y*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x_min, y
            index_pole_comp = x_max + (y*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x_min, y
        #ymin and ymax
        for x in range(x_min,x_max):
            index_pole_comp = x + (y_min*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x, y_min
            index_pole_comp = x + (y_max*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x, y_max
    
    #if nothing found in within range limit, return dummy number
    return 9000000,0
#==============================================================================


#==============================================================================
# Matching poles and connections by closest distance
def MatchPolesConn(indexes_conn,indexes_poles,d_EW_between,d_NS_between):
# This function creates an array to specify which poles connect to which houses, 
# by which houses are closest to which poles.    
    ConnPoles = np.zeros((len(indexes_conn[:,0]),2))      
    DistanceConnPoles = DistanceBWindexes(indexes_conn,indexes_poles,d_EW_between,d_NS_between)
    for i in range(len(ConnPoles[:,0])):
        ConnPoles[i,0] = np.argmin(DistanceConnPoles[i,:]) #The index of the closest pole for this connection, as indexed in "indexes_poles"
        ConnPoles[i,1] = np.min(DistanceConnPoles[i,:]) #The distance to the closest pole for this connection
    return ConnPoles
#==============================================================================

#===============================================================================
# Match Connections to Poles
def PolePlacement(reformatScaler,num_clusters,exclusionBuffer,range_limit,indexes_conn,indexes_excl,height,width,d_EW_between,d_NS_between):                    
# This function runs the previously mentioned functions together to determine 
# the final pole placements.       
    #Get initial solution of pole indexes with Gaussian Mean Clustering
    #The mean of the clusters is the initial pole placement
    ConnPole_initial, initial_pole_placement = Clustering(indexes_conn,num_clusters)
    indexes_poles = np.copy(initial_pole_placement)
    #ConnPole: the index is the connection # and the value is the pole
    #initial_pole_placement: the index is the pole #, the numbers are the indexes of that pole

    #Check if pole placed on exclusion
    #If in exclusion zone, circle outwardly to find non-exclusion spot
    #Turn 2D index arrays to 1D for comparison
    index_pole_comp = initial_pole_placement[:,0] + np.dot(initial_pole_placement[:,1],0.00001)
    index_excl_comp = indexes_excl[:,0] + np.dot(indexes_excl[:,1],0.00001)
    for i in range(len(index_pole_comp)):
        if index_pole_comp[i] in index_excl_comp:
            #check surrounding areas for non-exclusion spot
            index_x_new, index_y_new = FindNonExclusionSpot(initial_pole_placement[i,0], initial_pole_placement[i,1], index_excl_comp,range_limit,height,width)
            if index_x_new == 9000000: #arbitary large number to act as signal
                print("No new pole placement found for pole " +str(i)+", try expanding pole placement region.")
            else:
                #replace pole index location
                indexes_poles[i,:] = [index_x_new, index_y_new]
                #print("New pole placed at "+str(i)+" pole")
        
    #Match Connections and Poles  
    ConnPoles = MatchPolesConn(indexes_conn,indexes_poles,d_EW_between,d_NS_between)
    #Find in use poles and create new pole indexes
    indexes_poles_inuse_x = []
    indexes_poles_inuse_y = []
    for i in range(len(indexes_poles)):
        if i in ConnPoles[:,0]:
            indexes_poles_inuse_x.append(indexes_poles[i,0])
            indexes_poles_inuse_y.append(indexes_poles[i,1])
    indexes_poles_in_use_og = np.transpose(np.array([indexes_poles_inuse_x,indexes_poles_inuse_y]))
    #Redetermine match Connections and Poles  
    ConnPoles = MatchPolesConn(indexes_conn,indexes_poles_in_use_og,d_EW_between,d_NS_between)
       
    return ConnPoles, indexes_poles_in_use_og

#==============================================================================
             
 
#==============================================================================
# Calculate the Cost of the penalties to use as the minimizing optimization value
def PenaltiesToCost(reliability_cost, costs, dfpoles, dfnet, dfdropline):
    # This function calculates the total cost of the network including the equipment 
    # costs and reliability cost. The reliability cost can be set to zero.   
    net_cost = NetworkCost(costs, dfpoles, dfnet, dfdropline)
    
    Total_cost = net_cost['Line Total (USD)'].values.sum() + reliability_cost
    return Total_cost
#==============================================================================        
 
    
#==============================================================================
# Collect the connection and exclusions data
def CollectVillageData():
# This function collects all of the data of the community needed to determine the network layout.    
    
    #Load Files
    #Gather the information needed
    #Import csv file which has been converted from the klm file
    #This gives the points of connections which are houses to link to the distribution grid
    Connect_nodes = pd.read_excel('MAK_Connections.xlsx', sheet_name = 'sql_statement')
    Exclusion_nodes = pd.read_excel('MAK_exclusions.xlsx', sheet_name = 'MAK_exclusions')
            
    #Identify gps coordinate min and max to determine coordinates of edges of jpg image
    Longitude_exc = Exclusion_nodes['X']
    Latitude_exc = Exclusion_nodes['Y']
    #also convert these degrees to radians
    Lat_exc_min = m.radians(Latitude_exc.min()) #top of image (north)
    Lat_exc_max = m.radians(Latitude_exc.max()) #bottom of image (south)
    Long_exc_min = m.radians(Longitude_exc.min()) #left of image (east)
    Long_exc_max = m.radians(Longitude_exc.max()) #right of image (west)

    #Calculate the distance between the gps coordiantes using Haversine Formula
    #North South Distance #measuring latitude difference
    d_NS = GPStoDistance(Lat_exc_max,Lat_exc_min,Long_exc_max,Long_exc_max) #m
    #East West Distance #measuring longitude difference
    d_EW = GPStoDistance(Lat_exc_max,Lat_exc_max,Long_exc_max,Long_exc_min) #m

    #Import kml pdf file (of exclusions) and convert to jpg
    pages = convert_from_path('MAK_exclusions.pdf',500)
    for page in pages:
        page.save('MAK_exclusions.jpg','JPEG')
    
    #Convert JPG to array
    ExclusionMap = Image.open('MAK_exclusions.jpg')
    ExclusionMap_array = np.array(ExclusionMap)
    #Filter rgb value to 0 'non exclusion' and 1 'exclusion'
    #Black 0-0-0, White 255-255-255
    height = int(len(ExclusionMap_array[:,0])/reformatScaler) #this is y_index_max
    width = int(len(ExclusionMap_array[0,:])/reformatScaler) #this is x_index_max
    filename = "index_maxes_%s.csv" %str(reformatScaler)
    np.savetxt(filename,[height,width],delimiter=",")

    #Determine distance between reformatted pixels (between values in the array)
    d_EW_between = d_EW/width #m
    d_NS_between = d_NS/height #m
    filename = "d_between_%s.csv" %str(reformatScaler)
    np.savetxt(filename,[d_EW_between,d_NS_between],delimiter=",")

    #Load exlusion map, if not available then perform
    #This gathers the exclusion array indexes
    try:
        #print("in try loop")
        index_csv_name = "indexes_reformatted_%s_bufferzone_%s.csv" %(str(reformatScaler),str(exclusionBuffer))
        indexes_excl = np.loadtxt(index_csv_name, delimiter=",")
    except:
        print("in except loop")
        #quit()
        indexes_excl = ExclusionMapper(ExclusionMap_array,reformatScaler,exclusionBuffer,d_EW_between,d_NS_between,width,height)

    #Match the connection locations to locations in the array
    #Find distance between east limit of image and connection
    try:
        #print("in try loop")
        index_csv_name = "indexes_conn_reformatted_%s.csv" %str(reformatScaler)
        indexes_conn = np.loadtxt(index_csv_name, delimiter=",")
    except:
        d_Econnection = np.zeros(len(Connect_nodes))
        d_Nconnection = np.zeros(len(Connect_nodes))
        indexes_conn = np.zeros((len(Connect_nodes),2))
        for i in range(len(Connect_nodes)): #iteration through connections
            d_Econnection[i] = GPStoDistance(Lat_exc_min,Lat_exc_min,Long_exc_min,m.radians(Connect_nodes['x'][i])) #m
            #print(d_Econnection[i])
            #distance of connection to the east (left) (x index)
            d_Nconnection[i] = GPStoDistance(Lat_exc_min,m.radians(Connect_nodes['y'][i]),Long_exc_min,Long_exc_min) #m
            #print(d_Nconnection[i])
            #distance of connection to the north (top) (y index)
            #Get array index locations of all connections
            indexes_conn[i,0] = int(d_Econnection[i]/d_EW_between)
            #print(indexes_conn[i,0])
            indexes_conn[i,1] = int(d_Nconnection[i]/d_NS_between)
            #print(indexes_conn[i,1])
            index_csv_name = "indexes_conn_reformatted_%s.csv" %str(reformatScaler)
            np.savetxt(index_csv_name,indexes_conn, delimiter=",")
            
    return indexes_conn, indexes_excl, height, width, d_EW_between, d_NS_between,Long_exc_max, Long_exc_min,Lat_exc_max, Lat_exc_min

#==============================================================================
 
#==============================================================================
# Plot All Poles and All wiring
def Plot_AllPoles_AllWiring(POI, OnOff_MV, indexes_poles_MV, OnOff_groups, indexes_poles_groups, indexes_conn, ConnPoles, indexes_poles_LV_all):
# This function plots all the wiring (MV, LV, and house connections) together on the same plot. 
    
    #MV Poles and Wiring
    num_poles_MV = len(indexes_poles_MV[:,0])
    goodToGo = 0
    for i in range(num_poles_MV):
        for j in range(i):
            if OnOff_MV[i,j] == 1:
                match = np.array([[indexes_poles_MV[i,0],indexes_poles_MV[i,1]],[indexes_poles_MV[j,0],indexes_poles_MV[j,1]]]) 
                if goodToGo == 0:
                    goodToGo = 1
                    wiringMatrix_MV = np.copy(match)
                else:
                    wiringMatrix_MV = np.concatenate((wiringMatrix_MV,match),axis=1) 
    #MV Plot
    num_match_MV = len(wiringMatrix_MV[0,:])
    fig, ax = plt.subplots()
    for i in range(0,num_match_MV-1,2):
        j = i + 1
        plt.plot(wiringMatrix_MV[:,i],wiringMatrix_MV[:,j],color='green',linewidth=2)
    plt.scatter(indexes_poles_MV[:,0],indexes_poles_MV[:,1],s=2,c='b')
    #Plot POI with big black dot
    plt.scatter(POI[0],POI[1],s=5,c='black', marker='s')
    
    #LV Poles and Wiring
    cmap =  plt.cm.get_cmap('hsv', len(indexes_poles_groups))
    for group in range(len(indexes_poles_groups)):
        indexes_group = indexes_poles_groups[group]
        OnOff_group = OnOff_groups[group]
        num_poles_group = len(indexes_group)
        if len(indexes_group) != 1: #if only one index, skip (it will be under MV pole anyway)
            goodToGo = 0
            for i in range(num_poles_group):
                for j in range(i):
                    if OnOff_group[i,j] == 1:
                        match = np.array([[indexes_group[i][0],indexes_group[i][1]],[indexes_group[j][0],indexes_group[j][1]]]) 
                        if goodToGo == 0:
                            goodToGo = 1
                            wiringMatrix_group = np.copy(match)
                        else:
                            wiringMatrix_group = np.concatenate((wiringMatrix_group,match),axis=1) 
            #LV Plot
            num_match_group = len(wiringMatrix_group[0,:])
            for i in range(0,num_match_group-1,2):
                j = i + 1
                plt.plot(wiringMatrix_group[:,i],wiringMatrix_group[:,j],color = cmap(group),linewidth=1.5)
            for k in range(num_poles_group):
                plt.scatter(indexes_group[k][0],indexes_group[k][1],s=2,c='g')
            
        #Connections wiring
        num_conn = len(indexes_conn[:,0])
        for i in range(num_conn):
            pole = int(ConnPoles[i,0])
            x_s = [indexes_conn[i,0],indexes_poles_LV_all[pole,0]]
            y_s = [indexes_conn[i,1],indexes_poles_LV_all[pole,1]]
            plt.plot(x_s,y_s,color='orange',linewidth=1)
        
    ax.set_aspect('equal')
    plotname = "AllPolesAllWiring_Solution.png"
    plt.savefig(plotname, dpi=600)
    plt.show()
#==============================================================================
        
    
#==============================================================================
# Internal connected components recursive. This works because undirected
def Check_connections(i,visited,OnOff):
# This function is used internally to the ConnectedComponents and LineLosses 
# functions. It determines which lines have been visited relevant to those functions.    
    for j in range(len(OnOff[i,:])):
        if OnOff[i,j] == 1 and visited[j] == 0: #if j in a pole (999 is default for not a pole, and that pole is not visited)
            visited[j] = 1
            Check_connections(j,visited,OnOff)
    return visited
#==============================================================================
    
#==============================================================================
# Connected Components Code
def ConnectedComponents(OnOff):
# This function performed the graph theory connected components algorithm. It 
# checks whether all lines have been visited to determine is there are islands 
# in the network. 
    num_poles = len(OnOff[:,0])
    #Check # connected components
    visited = np.zeros(num_poles)
    conn_comp = 0
    for i in range(num_poles):
        if visited[i] == 0:
            visited[i] = 1
            conn_comp += 1
            visited = Check_connections(i,visited,OnOff)
        #else if visited continue to next pole
    
    return conn_comp, visited #return visited to varify everything has been visited
#==============================================================================

#==============================================================================
# Calc total load on poles
def PoleLoads(ConnPoles,num_poles):
    #ConnPoles is a single column numpy array
    num_conns = len(ConnPoles)
    load_per_connection = kW_max/num_conns
    PoleLoad_matrix = np.zeros(num_poles)
    for i in range(num_poles):
        PoleLoad_matrix[i] = int(np.ceil(np.count_nonzero(ConnPoles==i)))*load_per_connection        
    
    return PoleLoad_matrix
#==============================================================================
    
#==============================================================================
# Edges matrix: n-1 on line loss. |pole 1|pole 2|load loss|
def LineLosses(OnOff,ConnPoles,num_poles,Long_exc_min,Lat_exc_min,d_EW_between,d_NS_between,indexes_poles):
# This function calculates the load loss due to the loss of the lines in the network.   
    #POI of generation is always last pole in group for either MV or LV groups
    POI = len(OnOff[0,:])-1
    #Calculate load amount at each pole
    PoleLoad_matrix = PoleLoads(ConnPoles,num_poles)
    #Convert OnOff to list of edges (lines)
    num_edges = int(np.sum(OnOff)/2)
    Edges = np.zeros((num_edges,3))
    edge_count = 0
    for i in range(len(OnOff[:,0])):
        for j in range(i):
            if OnOff[i,j] == 1:
                OnOff_temp = np.copy(OnOff)
                OnOff_temp[i,j] = 0
                OnOff_temp[j,i] = 0
                #Calculate Load loss from this edge
                visited = np.zeros(num_poles)
                visited = Check_connections(POI,visited,OnOff_temp) #connected components to generation POI
                load_loss = 0
                for k in range(num_poles):
                    if visited[k] == 0:
                        load_loss = load_loss + PoleLoad_matrix[k]
                #Append Edges matrix with edge and load loss from loss of that edge (line)
                Edges[edge_count,:] = [i,j,load_loss]
                edge_count += 1
    return Edges
#============================================================================== 

#==============================================================================
# Wiring Algorithm
def WiringAlg(ConnPoles,prob,Cost_kWh,restoration_time,indexes_poles,wiring_cost,d_EW_between,d_NS_between,standalone):
# This function determines the lowest cost network layout using a network 
# reduction algorithm by using the previously mentioned functions.     
    t0 = time.time()

    #Load solution pole indexes
    if standalone == "True":
        filename = "indexes_poles_reformatted_%s_soln.csv" %(str(reformatScaler))
        indexes_poles = np.loadtxt(filename, delimiter=",")
        filename = "d_between_%s.csv" %str(reformatScaler)
        [d_EW_between,d_NS_between] = np.loadtxt(filename,delimiter=",")
    
    #Calculate Distances between all poles
    DistancesBWPoles = DistanceBWindexes(indexes_poles,indexes_poles,d_EW_between,d_NS_between)
    num_poles = len(indexes_poles[:,0])
             
    #Sort Distances between poles shortest connections to longest
    DistancesBWPoles_sorted = np.sort(DistancesBWPoles,axis=1) #First value in row will be 0, need to skip first column
    
    #Create Initial Solution
    goodToGo =  0
    num_conn_per_pole = 3 #starting value
    while goodToGo == 0:
        OnOff = np.zeros((num_poles,num_poles)) 
        for i in range(num_poles):
            for j in range(1,num_conn_per_pole): #start at 1 to avoid first column of zeros. Nothing is stopping this from going over num_poles
                ind = np.where(DistancesBWPoles == DistancesBWPoles_sorted[i,j]) #This is going to give two values, need adjust so only one half of matrix is considered
                #Place in use connection
                x = ind[0][0]
                y = ind[1][0]
                OnOff[x,y] = 1
                #make sure connection is also noted from the other pole's side (only will add up one side for total distances though so not double counting)
                OnOff[y,x] = 1
        # Check for islands
        components, visited = ConnectedComponents(OnOff) #need to make sure OnOff is being filled in way that there aren't rows of zeros 
        if components == 1:
            # Calculate Load loss risk
            Edges = LineLosses(OnOff,ConnPoles,num_poles,Long_exc_min,Lat_exc_min,d_EW_between,d_NS_between,indexes_poles)          
            total_load_loss_risk = sum(Edges[:,2])
            Reliability_Risk_Cost = total_load_loss_risk * prob * Cost_kWh * restoration_time
            #Solve for total distance and calc cost
            total_distance = 0
            for i in range(num_poles):
                for j in range(0,i):
                    if OnOff[i,j] == 1:
                        total_distance = total_distance + DistancesBWPoles[i,j]
            Wire_Cost = total_distance * wiring_cost
            Best_Total_Cost = Wire_Cost+Reliability_Risk_Cost #save as current best
            goodToGo = 1
        elif num_conn_per_pole != num_poles:
            num_conn_per_pole += 1
        else:
            OnOff = np.ones((num_poles,num_poles)) #set all connected except diagonal is 0
            for k in range(num_poles):
                OnOff[k,k] = 0
            # Calculate Load loss risk
            Edges = LineLosses(OnOff,ConnPoles,num_poles,Long_exc_min,Lat_exc_min,d_EW_between,d_NS_between,indexes_poles)          
            total_load_loss_risk = sum(Edges[:,2])
            Reliability_Risk_Cost = total_load_loss_risk * prob * Cost_kWh * restoration_time
            #Solve for total distance and calc cost
            total_distance = 0
            for i in range(num_poles):
                for j in range(0,i):
                    if OnOff[i,j] == 1:
                        total_distance = total_distance + DistancesBWPoles[i,j]
            Wire_Cost = total_distance * wiring_cost
            Best_Total_Cost = Wire_Cost+Reliability_Risk_Cost #save as current best
            goodToGo = 1
    Best_Reliability_Cost = np.copy(Reliability_Risk_Cost)
    Best_Wire_Cost = np.copy(Wire_Cost)
    Best_total_distance = np.copy(total_distance) 
    
    #Remove connections, starting with longest working towards shortest, check to make sure not creating islands
    DistancesBWPoles_in_use = DistancesBWPoles_sorted[:,1:num_conn_per_pole] #truncate
    DistancesBWPoles_in_use = DistancesBWPoles_in_use.flatten() #flatten
    DistancesBWPoles_in_use = np.sort(DistancesBWPoles_in_use) #sort
    for k in range(len(DistancesBWPoles_in_use)-1,-1,-1):
        OnOff_temp = np.copy(OnOff)
        ind = np.where(DistancesBWPoles == DistancesBWPoles_in_use[k])
        x = ind[0][0]
        y = ind[1][0]
        OnOff_temp[x,y] = 0
        OnOff_temp[y,x] = 0 #make other matching pair 0 as well
        # Check for islands
        components, visited = ConnectedComponents(OnOff_temp)
        if components == 1: #no islands, Only move forward with best solution if no islands
            # Calculate Load loss risk
            Edges = LineLosses(OnOff_temp,ConnPoles,num_poles,Long_exc_min,Lat_exc_min,d_EW_between,d_NS_between,indexes_poles)
            total_load_loss_risk = sum(Edges[:,2])
            Reliability_Risk_Cost = total_load_loss_risk * prob * Cost_kWh * restoration_time
            #Solve for total distance and calc cost
            total_distance = 0
            for i in range(num_poles):
                for j in range(i):
                    if OnOff_temp[i,j] == 1:
                        total_distance = total_distance + DistancesBWPoles[i,j]
            Wire_Cost = total_distance * wiring_cost
            Total_Cost = Wire_Cost+Reliability_Risk_Cost
            if Total_Cost < Best_Total_Cost: #Save best solution based on no islands and lowest cost
                OnOff = np.copy(OnOff_temp)
                Best_Total_Cost = np.copy(Total_Cost)
                Best_Reliability_Cost = np.copy(Reliability_Risk_Cost)
                Best_Wire_Cost = np.copy(Wire_Cost)
                Best_total_distance = np.copy(total_distance) 
        #print(Best_Total_Cost)
        #print(Best_Reliability_Cost)
        #print("********************")
                
    t1 = time.time()
    total_time = t1-t0
    
    return Best_total_distance, OnOff, DistancesBWPoles, num_conn_per_pole, Best_Reliability_Cost, Best_Wire_Cost, total_time
    
#==============================================================================

#==============================================================================
# Load behind each MV pole, to check that not over 220V line limit
def loadBehindPoles(ConnPoles_MV, ConnPoles_LV, num_conns,MV_pole_num,load_per_conn):
# This function determines the amount of load behind each pole to determine if 
# the connected line is rated to handle that load capacity.     
    ConnPoles_LVMV = np.zeros(num_conns) #the MV pole that a house connects to
    load_behind_MV_poles = np.zeros(MV_pole_num)
    for i in range(num_conns):
        Conn = int(ConnPoles_LV[i,0])
        ConnPoles_LVMV[i] = ConnPoles_MV[Conn,0]
    for j in range(MV_pole_num):
        load_behind_MV_poles[j] = np.count_nonzero(ConnPoles_LVMV==j)*load_per_conn
    max_load_behind = max(load_behind_MV_poles)
    return max_load_behind, ConnPoles_LVMV
#==============================================================================        
    
#==============================================================================
# Calculate Closest Pole to POI (point of interconnection) to generation
def POI_Pole(lat_Generation,long_Generation,Long_exc_min,Lat_exc_min,d_EW_between,d_NS_between,indexes_poles,d_BW_Adj_Poles):
    EW_dis = GPStoDistance(Lat_exc_min,Lat_exc_min,Long_exc_min,long_Generation)
    NS_dis = GPStoDistance(Lat_exc_min,lat_Generation,Long_exc_min,Long_exc_min)
    EW_index = int(EW_dis/d_EW_between)
    NS_index = int(NS_dis/d_NS_between)
    indexes_gen = [EW_index,NS_index]
    #Calculate distances between generation and pole
    #Individually feed in each pair
    num_poles = len(indexes_poles[:,0])
    Distance_Gen_Poles = np.zeros(num_poles)    
    for i in range(num_poles):
        Distance = DistanceBWindexes(indexes_gen,indexes_poles[i,:],d_EW_between,d_NS_between) #Type error: only size-1 arrays can be converted to Python Scalars
        Distance_Gen_Poles[i] = Distance
    closest_pole = np.argmin(Distance_Gen_Poles)
    Distance_Gen_Poles[closest_pole] = 90000000000 # put dummy number
    closest_pole2 = np.argmin(Distance_Gen_Poles)
    edge_dist = d_BW_Adj_Poles/np.sqrt(d_EW_between**2 + d_NS_between**2)
    edge = [indexes_poles[closest_pole, :], indexes_poles[closest_pole2, :],
            np.sqrt(np.sum(np.power(indexes_poles[closest_pole2, :] - indexes_poles[closest_pole, :], 2)))]
    
    Points = IntermediatePoints(edge, edge_dist)
    new_distances = np.zeros(len(Points))
    for i in range(len(Points)):
        new_distances[i] =  DistanceBWindexes(indexes_gen, Points[i, :],d_EW_between,d_NS_between)
    idx = np.argmin(new_distances)
    return [closest_pole, closest_pole2], indexes_gen, Points[idx, :]
#==============================================================================


#==============================================================================
# Given an edge (p1, p2, d_p1_p2), and allowable distance between adjacent points
# create intermediate points along the straight line connecting the two nodes
def IntermediatePoints(edge, edge_dist):
    P1, P2, k = edge # unpack points
    if k < 1.3*edge_dist:
        return np.vstack((np.array([P1[0], P1[1]]), np.array([P2[0], P2[1]])))

    nodes_num = int(k//edge_dist) # number of nodes to add
    
    if nodes_num == 0:
        nodes_num = 1
        
    h_incr = 0 # horizontal increment
    v_incr = 0 # vertical increment
    if P2[0] == P1[0]:
        if P2[1] > P1[1]:
            theta = np.pi/2
            v_incr = edge_dist
        else:
            theta = 3*np.pi/2
            v_incr = -edge_dist
    if P2[1] == P1[1]:
        if P2[0] > P1[0]:
            theta = 0
            h_incr = edge_dist
        else:
            theta = np.pi
            h_incr = -edge_dist
    if P2[0] != P1[0] and P2[1] != P1[1]:
        g = (P2[1] - P1[1])/(P2[0] - P1[0]) # evaluate gradient
        # Determine theta and direction
        theta = m.atan(g)
        if P2[0] > P1[0] and P2[1] > P1[1]: # 1st Quadrant 
            h_incr = edge_dist*m.cos(theta) 
            v_incr = edge_dist*m.sin(theta) 
            
        elif P2[0] < P1[0] and P2[1] > P1[1]: # 2nd Quadrant 
             h_incr = -edge_dist*m.cos(theta) 
             v_incr = -edge_dist*m.sin(theta) 
             
        elif P2[0] < P1[0] and P2[1] < P1[1]: # 3rd Quadrant 
             h_incr = -edge_dist*m.cos(theta) 
             v_incr = -edge_dist*m.sin(theta) 
             
        elif P2[0] > P1[0] and P2[1] < P1[1]: # 4th Quadrant 
             h_incr = edge_dist*m.cos(theta) 
             v_incr = edge_dist*m.sin(theta)
    
    new_nodes = np.array([P1[0], P1[1]])
    for i in range(nodes_num):
        x_i = P1[0] + (i + 1)*h_incr
        y_i = P1[1] + (i + 1)*v_incr
        node = np.array([int(x_i), int(y_i)])
        new_nodes = np.vstack((new_nodes, node))
    new_nodes = np.vstack((new_nodes, np.array([P2[0], P2[1]])))
    
    # If the distance between the last two poles is less than 30%, remove the 
    # last intermediate pole
    d_last = np.sqrt(np.sum(np.power(new_nodes[-1, :] - new_nodes[-2,:], 2)))
    if d_last < 0.15*edge_dist:
        new_nodes = np.vstack((new_nodes[:-3, :], new_nodes[-1, :]))
    return new_nodes
#==============================================================================

#==============================================================================
# Evaluate Intermediate Poles respecting exclusion zones
def MVIntermediatePoles(indexes, OnOffMV, d_BW_Poles, d_BW_Adj_Poles, index_excl_comp, range_limit, max_y, max_x):
    
    OnOff = np.triu(OnOffMV) # Take connectivity from the upper triangle (above diagonal)
    # Taking the lower triangle would reverse starting points and end points
    # and would have to be indexed by columns instead of rows
    pairs = []
    for i, index in enumerate(indexes):
        idxs = np.where(OnOff[i, :] == 1)[0]
        if not list(idxs):
            pass
        else:
            for idx in idxs:
                d = d_BW_Poles[i][idx]
                pairs.append([index, indexes[idx, :], d])
                
    # Convert absolute distance to pixel distance
    edge = pairs[0]
    edge_d = np.sqrt(np.sum((edge[0] - edge[1])**2))
    pixel_d = edge_d/edge[2] # pixel distance per meter
    threshold = d_BW_Adj_Poles*pixel_d # 
    
    new_pairs = np.array([])
    for e in pairs:
        e[2] = e[2]*pixel_d
        if not list(new_pairs):
            new_pairs = IntermediatePoints(e, threshold)
        else:
            new_pairs = np.vstack((new_pairs, IntermediatePoints(e, threshold)))
    new_pairs = np.unique(new_pairs, axis=0) # Remove repeating points
    
    # Respect the exclusion zones
    for i, pt in enumerate(new_pairs):
        if pt[1] in indexes[:, 1] and pt[0] in indexes[:, 0]:
            pass
        else:
            x, y = FindNonExclusionSpot(pt[0], pt[1], index_excl_comp, range_limit, max_y, max_x)
            new_pairs[i, :] = np.array([x, y])
        
    return  new_pairs
#==============================================================================


#==============================================================================
# Evaluate angle formed by three points
def AngleBWPoints(P_END_1, P_Mid, P_END_2, d_EW_between, d_NS_between):
    
    def angle(P1, P2):
        if P2[0] == P1[0] and P2[1] > P1[1]:
            return np.pi/2
        elif P2[0] == P1[0] and P2[1] < P1[1]:
            return 3*np.pi/2
        elif P2[1] == P1[1] and P2[0] > P1[0]:
            return 0
        elif P2[1] == P1[1] and P2[0] < P1[0]:
            return np.pi
        else:
            g = (P1[1] - P2[1])*d_NS_between/((P1[0] - P2[0])*d_EW_between)
            theta = m.atan(g)
            
            if P2[1] > P1[1] and P2[0] > P1[0]: # First Quadrant
                return theta
            elif P2[1] > P1[1] and P2[0] < P1[0]: # Second Qudrant
                return np.pi - np.abs(theta)
            elif P2[1] < P1[1] and P2[0] < P1[0]: # Third Quadrant
                return theta + np.pi
            elif P2[1] < P1[1] and P2[0] > P1[0]: # Fourth Quadrant
                return 2*np.pi - np.abs(theta)
    theta1 = angle(P_Mid, P_END_1)
    theta2 = angle(P_Mid, P_END_2)
    angles = np.degrees([theta1, theta2])
    if abs(angles[1] - angles[0]) <= 180:
        return 180 - abs(angles[1] - angles[0])
    else: 
        return 180 - (360 - abs(angles[1] - angles[0]))
    
#==============================================================================



#==============================================================================
# Return Minimum Spanning Tree 
def MST(indexes, d_EW_between, d_NS_between):
    """
    This function implements a least cost network using Prim's minimum spanning 
    tree algorithm
    """
    points = list(zip(indexes[:, 0]*d_EW_between, indexes[:, 1]*d_NS_between))
    try:
        rand_point = np.random.choice(points)
    except:
        rand_point = points[0]
    fPts = [rand_point] # initialize arbitrary starting points
    pts = list(points.copy()) # create shallow copy of the points
    pts.remove(rand_point)
    cPts = [] # connected points
    while len(pts) != 0:
        dist = [] # distances to all points
        dmins = [] # minimum distances
        for pt in fPts:
            dist.append([((pt[0] - pti[0])**2 + (pt[1] - pti[1])**2)**(1/2) for pti in pts])
        for dlist in dist:
            dmins.append(min(dlist)) # add minimum distance to this list
        # determine the index of the list with minimum distance    
        list_idx = dmins.index(min(dmins)) 
        # find the index of the minimum distance in a list it is in
        mind_idx = dist[list_idx].index(min(dist[list_idx]))
        
        cPts.append([fPts[list_idx], pts[mind_idx]]) # add a pair of connected points
        fPts.append(pts[mind_idx])
        pts.remove(pts[mind_idx])
    G = nx.Graph()
    G.add_edges_from(cPts)
    return G
#==============================================================================

#==============================================================================
# Return Minimum Spanning Tree Connectivity matrix
def MSTConnectivityMatrix(indexes, d_EW_between, d_NS_between):
    points = list(zip(indexes[:, 0]*d_EW_between, indexes[:, 1]*d_NS_between))
    G = MST(indexes, d_EW_between, d_NS_between)
    connectivityMatrix = nx.adjacency_matrix(G, nodelist=points).todense()
    return connectivityMatrix
#==============================================================================


#==============================================================================
# Put poles into respective branches
def EvaluateBranches(source_index, indexes, d_EW_between, d_NS_between):
    source_node = (source_index[0]*d_EW_between, source_index[1]*d_NS_between) 
    G = MST(indexes, d_EW_between, d_NS_between)
    G.remove_node(source_node)
    branches = list(nx.connected_components(G))
    branches_ = []
    for b in branches:
        indexes_ = np.array(list(b))
        indexes_[:, 0] = indexes_[:, 0]/d_EW_between
        indexes_[:, 1] = indexes_[:, 1]/d_NS_between
        branches_.append(indexes_)
    return branches_
#==============================================================================


#==============================================================================
# Classify poles by angle
def PoleAngleClass(pole_indexes, d_EW_between, d_NS_between, ntype):
    connectivity = MSTConnectivityMatrix(pole_indexes, d_EW_between, d_NS_between)
    classes = []
    for i in range(len(pole_indexes)):
        neighbors = np.sum(connectivity[i, :])
        p_mid = pole_indexes[i, :]
        if neighbors == 1:
            classes.append('terminal')
        elif neighbors == 2:
            neighbors_idx = np.where(connectivity[i, :] == 1)[1]
            p_end_1 = pole_indexes[neighbors_idx[0], :]
            p_end_2 = pole_indexes[neighbors_idx[1], :]
            agl = AngleBWPoints(p_end_1, p_mid, p_end_2, d_EW_between, d_NS_between)
            if ntype == 'LV':
                if agl == 0 or agl == 180:
                    classes.append("mid_straight")
                elif agl < 60:
                    classes.append("mid_less_60")
                elif agl > 60:
                    classes.append("mid_over_60")
            elif ntype == 'MV':
                if agl == 0 or agl == 180:
                    classes.append("mid_straight")
                elif agl < 30:
                    classes.append("mid_less_30")
                elif agl > 30:
                    classes.append("mid_over_30")
        elif neighbors > 2:
            neighbors_idx = np.where(connectivity[i, :] == 1)[1]
            p_ends = pole_indexes[neighbors_idx, :]
           
            d_BW_p_ends = DistanceBWindexes(p_ends, p_ends, d_EW_between, d_NS_between)
            d_BW_p_ends = d_BW_p_ends + np.diagflat(np.ones(neighbors)*99999999)
            # Evaluate for all combinations of middle pole and all neighbors
            angles = np.zeros(neighbors)
            for j in range(neighbors):
                n1 = p_ends[j, :]
                index_n2 = np.argmin(d_BW_p_ends[j, :])
                n2 = p_ends[index_n2, :]
                d_BW_p_ends[index_n2, j] = 999999
                angles[j] = AngleBWPoints(n1, p_mid, n2, d_EW_between, d_NS_between)
                
            agl = angles.min()
            if ntype == 'LV':
                if agl == 0 or agl == 180:
                    classes.append("mid_straight")
                elif agl < 60:
                    classes.append("mid_less_60")
                elif agl > 60:
                    classes.append("mid_over_60")
            elif ntype == 'MV':
                if agl == 0 or agl == 180:
                    classes.append("mid_straight")
                elif agl < 30:
                    classes.append("mid_less_30")
                elif agl > 30:
                    classes.append("mid_over_30")
    classes_ = pd.DataFrame(pole_indexes, columns = ['index_x', 'index_y'])
    classes_['AngleClass'] = classes 
    return classes_
#==============================================================================


#==============================================================================
# Get elevation of pole indexes using the generation latitude and logitude as 
# reference point from the open access API open-elevation
def PoleElevation(gen_LON, gen_LAT, gen_indexes, target_indexes, d_EW_between, d_NS_between):
    import requests
    GPD = gpd.GeoDataFrame(geometry=[Point(np.degrees(gen_LAT), np.degrees(gen_LON))])
    GPD = GPD.set_crs("EPSG:4326") # WGS84
    GPD = GPD.to_crs("EPSG:32733") # Transform to UTM (Mercator)
    X, Y = list(GPD.geometry.values[0].coords)[0]
    X_Shifts = (gen_indexes[0] - target_indexes[:, 0])*d_EW_between
    Y_Shifts = (gen_indexes[1] - target_indexes[:, 1])*d_NS_between
    DF = pd.DataFrame({'index_x':target_indexes[:,0], 'index_y':target_indexes[:,1],
                       'UTM_X':X - X_Shifts, 'UTM_Y':Y - Y_Shifts})
    new_GDF = gpd.GeoDataFrame(DF, geometry=gpd.points_from_xy(DF.UTM_X, DF.UTM_Y))
    new_GDF = new_GDF.set_crs("EPSG:32733")
    new_GDF = new_GDF.to_crs("EPSG:4326") 
    gps = np.array([[list(i.coords)[0][0], list(i.coords)[0][1]] for i in new_GDF.geometry.values])
    url = "https://maps.googleapis.com/maps/api/elevation/json?locations="
    for idx, pt in enumerate(new_GDF.geometry.values):
        lat, lon = list(pt.coords)[0]
        if idx < len(new_GDF) - 1:
            url = url + str(lat) + ',' + str(lon) + '|'
        else:
            url = url + str(lat) + ',' + str(lon)
    url = url + '&key=AIzaSyB7NCOraSbaIBDVg2-BU5D_mX_Q2BwZV2E'
    payload = {}
    headers = {}
    response = requests.request('GET', url, headers=headers, data=payload).json()
    elevations = [res['elevation'] for res in response['results']]
    return np.array(elevations),  list(DF.UTM_X), list(DF.UTM_Y), gps
#==============================================================================


#==============================================================================
# Evaluate the altitude angle
def CorrectLengthAngle(horizontald1_2, elevation1, elevation2):
    net_elevation = np.abs(elevation1 - elevation2)
    cosine = horizontald1_2/np.sqrt(net_elevation**2 + horizontald1_2**2)
    return [np.degrees(m.acos(cos_)) for cos_ in cosine]
#==============================================================================


#==============================================================================
# Evaluate network segments length on a 2D plane
def NetworkLength(indexes, d_EW_between, d_NS_between):
    mst = MST(indexes, d_EW_between, d_NS_between)
    edges =list(mst.edges)
    og_indexes = np.zeros((len(edges), 4))
    for i, e in enumerate(edges):
        X = np.array(e)
        og_indexes[i, 0] = X[0, 0]/d_EW_between
        og_indexes[i, 1] = X[0, 1]/d_NS_between
        og_indexes[i, 2] =  X[1, 0]/d_EW_between
        og_indexes[i, 3] = X[1, 1]/d_NS_between
    df = pd.DataFrame(data=og_indexes, columns=['index_x_from', 'index_y_from', 
                                                'index_x_to', 'index_y_to'])
    gdf = gpd.GeoDataFrame(df, geometry=[LineString(e) for e in edges])
    gdf['length'] = [s.length for s in gdf.geometry.values]
    return gdf
#==============================================================================


#==============================================================================
# Evaluate the cost of the network
def NetworkCost(costs, dfpoles, dfnet, dfdropline):
    result_df = pd.DataFrame()
    references = ['Assembly - Pole - MV - Start',
                  'Assembly - Step - Down - Transformer',
                  'Assembly - Pole - MV - Mid',
                  'Assembly - Pole - MV - Bend <30',
                  'Assembly - Pole - MV - Bend >30',
                  'Assembly - Pole - MV - End',
                  'Assembly - Pole - LV - Mid',
                  'Assembly - Pole - LV - Bend <60',
                  'Assembly - Pole - LV - Bend >60',
                  'Assembly - Pole - LV - End',
                  'Wire - MV',
                  'Wire - LV',
                  'Wire - LineDrop',
                  'Assembly - Meter']
   
    mv_ref = dfpoles[dfpoles.Type == 'MV']
    mv_ref = mv_ref[mv_ref.distance_from_source != 0]
    lv_ref = dfpoles[dfpoles.Type == 'LV']
    quantity = np.zeros(len(references))
    quantity[0] = 1
    num_transformers = 0
    for item in mv_ref.ID.values:
        if item[-1].isalpha() == True:
            num_transformers += 1
    quantity[1] = num_transformers
    
    mv_pole_counts = mv_ref.value_counts(subset=['AngleClass'])
    try:
        quantity[2] = mv_pole_counts['mid_straight']
    except:
        pass
    try:
        quantity[3] = mv_pole_counts['mid_less_30']
    except:
        pass
    try:
        quantity[4] = mv_pole_counts['mid_over_30']
    except:
        pass
    try:
        quantity[5] = mv_pole_counts['terminal']
    except:
        pass
    
    lv_pole_counts = lv_ref.value_counts(subset=['AngleClass'])
    try:
        quantity[6] = lv_pole_counts['mid_straight']
    except:
        pass
    try:
        quantity[7] = lv_pole_counts['mid_less_60']
    except:
        pass
    try:
        quantity[8] = lv_pole_counts['mid_over_60']
    except:
        pass
    try:
        quantity[9] = lv_pole_counts['terminal']
    except:
        pass
    
    # Lines length
    mvnet = dfnet[dfnet.Type == 'MV']
    lvnet = dfnet[dfnet.Type != 'MV']
    quantity[10] = mvnet.adj_length.values.sum()
    quantity[11] = lvnet.adj_length.values.sum()
    
    # Line drop to households
    quantity[12] = dfdropline.Linedrop.values.sum()
    quantity[13] = len(dfdropline)
    
    temp_df = pd.DataFrame({'Part Reference':references,
                            'Qty':quantity,
                            'Price (USD)':costs})
    
    result_df = temp_df[['Part Reference', 'Qty','Price (USD)']]
    result_df['Line Total (USD)'] = result_df['Qty']*result_df['Price (USD)']
    return result_df
#==============================================================================


#==============================================================================
# Classify the poles by concession, MV/LV, Angle 
def ClassifyNetworkPoles(concession, gen_LAT, gen_LON, gen_site_indexes, indexes_MV_Poles_wPOI, group_indexes_LV,
                         OnOff_MV, d_BW_Poles_MV, index_excl_comp, d_BW_Adj_Poles, range_limit, max_y, max_x,
                         d_EW_between, d_NS_between, indexes_conn, conc_id = None):
    base = concession.upper()[0:3]
    if conc_id != None:
        base = concession.upper()[0:3] + str(conc_id)
    # MV Pole  and LV Pole Naming
    MV_poles_names = []
    MV_Pole_indexes = np.array([])
    
    LV_Poles_names = []
    LV_Pole_indexes = np.array([])
    
    AngleClasses = pd.DataFrame()
    network_length = gpd.GeoDataFrame()
    droplines = pd.DataFrame()
    basen = base + 'M'
    gen_index = gen_site_indexes
    if not list(MV_Pole_indexes):
        MV_Pole_indexes = np.array([gen_index[0], gen_index[1], 0])
    else:
        MV_Pole_indexes = np.vstack((MV_Pole_indexes, np.array([gen_index[0], gen_index[1], 0])))
    MV_poles_names.append(basen + str(1))
    t_gen_idx = np.array([gen_index[0]*d_EW_between, gen_index[1]*d_NS_between])
    mv_poles = indexes_MV_Poles_wPOI
    Yt = mv_poles*np.array([d_EW_between, d_NS_between])
    ydists = np.sqrt(np.sum(np.power(Yt - t_gen_idx, 2), axis=1))
    ydists_copy = np.copy(ydists)
    ydists_copy.sort()
    A = np.zeros((len(mv_poles), 2))
    ordered_idxs = [np.where(ydists == val)[0][0] for val in ydists_copy]
    A[:, 0] = mv_poles[:, 0][ordered_idxs]
    A[:, 1] = mv_poles[:, 1][ordered_idxs]
    all_MV_Poles = MVIntermediatePoles(mv_poles, OnOff_MV, d_BW_Poles_MV, d_BW_Adj_Poles, index_excl_comp, range_limit, max_y, max_x)
    angle_class_ = PoleAngleClass(all_MV_Poles, d_EW_between, d_NS_between, 'MV')
    if len(AngleClasses) == 0:
        AngleClasses = angle_class_
    else:
        AngleClasses = pd.concat([AngleClasses, angle_class_])
    net_len = NetworkLength(all_MV_Poles, d_EW_between, d_NS_between)
    net_len['Type'] = 'MV'
    if len(network_length) == 0:
        network_length = net_len
    else:
        network_length = pd.concat([network_length, net_len])
    branches = EvaluateBranches(gen_index, all_MV_Poles, d_EW_between, d_NS_between)
    for j, branch in enumerate(branches):
        n = j*100
        Xt = branch*np.array([d_EW_between, d_NS_between])
        dists = np.sqrt(np.sum(np.power(Xt - t_gen_idx, 2), axis=1))
        dists_copy = np.copy(dists)
        dists_copy.sort()
        B = np.zeros((len(branch), 3))
        ordered_index = [np.where(dists == val)[0][0] for val in dists_copy]
        B[:, 0:2] = branch[ordered_index]
        B[:, 2] = dists_copy
        MV_Pole_indexes = np.vstack((MV_Pole_indexes, B))
        names_ = [basen + str(n + i) for i in range(2, len(branch)+2)]
        t_counter = 65 # why start at 65? letter A is represented by 65, B 66 and so on
        for k in range(len(A)):
            if A[k, :] in branch:
                pos_row, pos_col = np.where(B[:, 0:2] == A[k, :])
                names_[pos_row[0]] = names_[pos_row[0]] + str(chr(t_counter))
                t_counter += 1
        MV_poles_names += names_
    
    # Low Voltage Poles
    group = group_indexes_LV
    sub_network_counter = 65
    for id_, g in enumerate(group):
        net_len = NetworkLength(g, d_EW_between, d_NS_between)
        net_len['Type'] = 'LV' + str(id_ + 1)
        if len(network_length) == 0:
            network_length = net_len
        else:
            network_length = pd.concat([network_length, net_len])
        branch_counter = 65
        for a in A:
            del_idx_x, del_idx_y = np.where(g == a)
            if len(del_idx_x) == 2 and del_idx_x[0] == del_idx_x[1]:
            #if a in g:
                temp_pole_class = PoleAngleClass(g, d_EW_between, d_NS_between, 'LV')
                #temp_pole_class.remove(temp_pole_class[del_idx_x[0]])
                if len(AngleClasses) == 0:
                    AngleClasses = temp_pole_class
                else:
                    AngleClasses = pd.concat([AngleClasses, temp_pole_class])
                branches_ = EvaluateBranches(a, g, d_EW_between, d_NS_between)
                for b_ in branches_:
                    source = np.array([a[0]*d_EW_between, a[1]*d_NS_between])
                    X = b_*np.array([d_EW_between, d_NS_between])
                    dists = np.sqrt(np.sum(np.power(X - source, 2), axis=1))
                    dists_copy = np.copy(dists)
                    dists_copy.sort()
                    C = np.zeros((len(b_), 3))
                    ordered_index = [np.where(dists == val)[0][0] for val in dists_copy]
                    C[:, 0:2] = b_[ordered_index]
                    C[:, 2] = dists_copy
                    if not list(LV_Pole_indexes):
                        LV_Pole_indexes = C
                    else:
                        LV_Pole_indexes = np.vstack((LV_Pole_indexes, C))
                    LV_Poles_names += [base + chr(sub_network_counter) + chr(branch_counter) + str(m+1) for m in range(len(b_))]
                branch_counter += 1
        sub_network_counter += 1
        
    # Determine customer connections
    lv_poles = np.concatenate(group)
    del_idx_x, del_idx_y = [], []
    for ir in range(len(mv_poles)):
        x = mv_poles[ir, 0]
        y = mv_poles[ir, 1]
        idx_x = np.where(lv_poles[:, 0] == x)[0]
        idx_y = np.where(lv_poles[:, 1] == y)[0]
        if not list(idx_x) or not list(idx_y):
            pass
        elif idx_x[0] == idx_y[0]:
            del_idx_x.append(idx_x[0])
            del_idx_y.append(idx_y[0])
    lv_poles_ = np.delete(lv_poles, del_idx_x, axis=0)
    
    dist_poles_conns = DistanceBWindexes(lv_poles_, indexes_conn, d_EW_between, d_NS_between)
    min_dists = dist_poles_conns.min(axis=0)
    min_idxs_x, min_idxs_y = np.where(dist_poles_conns == min_dists)
    T_lv_poles = lv_poles_*np.array([d_EW_between, d_NS_between])
    T_associated_conns = indexes_conn*np.array([d_EW_between, d_NS_between])
    temp_from_idx, temp_to_idx, temp_from_idy, temp_to_idy = [],[],[],[]
    lines = []
    for idx, idy in zip(min_idxs_x, min_idxs_y):
        temp_from_idx.append(lv_poles_[idx, 0])
        temp_to_idx.append(indexes_conn[idy, 0])
        temp_from_idy.append(lv_poles_[idx, 1])
        temp_to_idy.append(indexes_conn[idy, 1])
        lines.append(LineString((T_lv_poles[idx, :], T_associated_conns[idy, :])))
    temp_df = pd.DataFrame({'index_x_from':temp_from_idx,
                            'index_y_from':temp_from_idy,
                            'index_x_to':temp_to_idx,
                            'index_y_to':temp_to_idy,
                            'Linedrop':min_dists})
    droplines = gpd.GeoDataFrame(temp_df, geometry=lines)
    
    # Put pole classes into a dataframe and geodataframe
    DF = pd.DataFrame(data=np.vstack((MV_Pole_indexes, LV_Pole_indexes)), columns=['index_x', 'index_y', 'distance_from_source'])
    DF['ID'] = MV_poles_names + LV_Poles_names
    DF['Type'] = ['MV']*len(MV_poles_names) + ['LV']*len(LV_Poles_names)
    AngleClasses = AngleClasses.drop_duplicates(subset=['index_x', 'index_y'])
    
    DF = DF.sort_values(by=['index_x', 'index_y'])
    AngleClasses = AngleClasses.sort_values(by=['index_x', 'index_y'])
    DF['AngleClass'] = AngleClasses['AngleClass'].values
    
    gen_index = gen_site_indexes
    data = DF#[DF.concession == conc]
    target_indexes = data.filter(items=['index_x', 'index_y']).to_numpy()
    el, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_site_indexes, target_indexes, d_EW_between, d_NS_between)
    DF['elevation'] = el
    DF['UTM_X'] = utm_x
    DF['UTM_Y'] = utm_y
    DF['GPS_X'] = list(gps[:, 0])
    DF['GPS_Y'] = list(gps[:, 1])
    GDF = gpd.GeoDataFrame(DF, geometry=gpd.points_from_xy(DF.index_x*d_EW_between,
                                                           DF.index_y*d_NS_between))
    
    # Put nework lines into a dataframe and geodataframe
    from_indexes = np.zeros((len(network_length), 2))
    from_indexes[:, 0] = network_length.index_x_from.values
    from_indexes[:, 1] = network_length.index_y_from.values
    
    to_indexes = np.zeros((len(network_length), 2))
    to_indexes[:, 0] = network_length.index_x_to.values
    to_indexes[:, 1] = network_length.index_y_to.values
    el1, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_site_indexes, from_indexes, d_EW_between, d_NS_between)
    el2, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_site_indexes, to_indexes, d_EW_between, d_NS_between)
    pole_elevation = CorrectLengthAngle(network_length.length.values, el1, el2)
    network_length['elevation_angle'] = pole_elevation
    network_length['adj_length'] = network_length['length']/np.cos(np.radians(network_length.elevation_angle))
    
    connections = pd.DataFrame({'index_x':indexes_conn[:, 0], 'index_y':indexes_conn[:, 1]})
    geometry = [Point(p[0], p[1]) for p in T_associated_conns]
    connections = gpd.GeoDataFrame(connections, geometry=geometry)
    
    return GDF, network_length, droplines, connections
#==============================================================================


#==============================================================================
# Put the results in an excel file 
def ConcessionDetails(dfpoles, dfnet, dfdropline, dfcosts, connections, concession, conc_id=None):
    path = os.path.join(os.getcwd(), concession) # directory to save in
    if os.path.isdir(path):
        pass
    else:
        os.mkdir(path)
    # Save these into excel
    wb = Workbook()
    wb.create_sheet(title="PoleClasses", index=0)
    wb.create_sheet(title="NetworkLength", index=1)
    wb.create_sheet(title="DropLines", index=2)
    wb.create_sheet(title="Connections", index=3)
    wb.create_sheet(title="NetworkLayout", index=4)
    wb.create_sheet(title='NetworkCost', index=5)
    ws = wb["PoleClasses"]
    for row in dataframe_to_rows(dfpoles.drop(columns=['geometry']), header=True):
        ws.append(row)
    ws = wb["NetworkLength"]
    for row in dataframe_to_rows(dfnet.drop(columns=['geometry']), header=True):
        ws.append(row)
    ws = wb['DropLines']
    for row in dataframe_to_rows(dfdropline.drop(columns=['geometry']), header=True):
        ws.append(row)
    ws = wb['Connections']
    for row in dataframe_to_rows(connections.drop(columns=['geometry']), header=True):
        ws.append(row)
    
    # Evaluate date and time on which the simulation is done
    simdate = dt.datetime.today() # simulation date

    # Define a function that ensures double digits for day of the month
    add0 = lambda x: '0'+str(x) if x < 10 else str(x)
    
    filename = str(simdate.year) + add0(simdate.month) + add0(simdate.day) + \
        '_' + add0(simdate.hour) + add0(simdate.minute) + '_' + concession.upper()[0:3] + \
        '_uGrid'
    ws = wb["NetworkLayout"]
    fig, ax = plt.subplots(dpi=250)
    dfpoles[dfpoles.Type == "MV"].plot(column='Type', s=8, ax=ax, legend=False)
    dfpoles[dfpoles.Type == "LV"].plot(column='Type', s=4, ax=ax, legend=False)
    connections.plot(markersize=2, color='brown', ax=ax, legend=False)
    dfdropline.plot(lw=0.5, ax=ax, color='red', legend=False)
    dfnet[dfnet.Type != "MV"].plot(column="Type", lw=0.7,ax=ax,legend=False)
    dfnet[dfnet.Type == "MV"].plot(lw=1,ax=ax, color='blue',legend=False)
    plt.savefig(concession+'.png')
    ws.add_image(img(concession+'.png'), 'B4')
    
    ws = wb['NetworkCost']
    for row in dataframe_to_rows(dfcosts, index=True, header=True):
        ws.append(row)
    free_cell_loc = len(dfcosts)+3
    ws['D'+str(free_cell_loc)] = 'Total'
    ws['E'+str(free_cell_loc)] = dfcosts['Line Total (USD)'].values.sum()
    if conc_id == None:
        wb.save(path+'/'+ filename + ".xlsx")
    else:
        wb.save(path+'/'+ filename + str(conc_id)+ ".xlsx")
#==============================================================================

#==============================================================================
def AddSpur(concession, filename, gen_LON, gen_LAT, gen_indexes, indexes, type_,
            d_EW_between, d_NS_between, costs, conc_id=None):
    path = os.path.join(os.getcwd(), concession) # directory
    try:
        poleclasses = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='PoleClasses', skiprows=[1])
        networklength = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='NetworkLength', skiprows=[1])
        droplines = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='DropLines', skiprows=[1])
        conns = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='Connections', skiprows=[1])
    except:
        poleclasses = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='PoleClasses', skiprows=[1])
        networklength = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='NetworkLength', skiprows=[1])
        droplines = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='DropLines', skiprows=[1])
        conns = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='Connections', skiprows=[1])
    poleclasses = poleclasses.drop(columns= ['Unnamed: 0'])
    pclasses = poleclasses[poleclasses.Type == type_.upper()]
    
    pole_indexes = pclasses.filter(items=['index_x', 'index_y']).to_numpy()
    pole_ids = pclasses.ID.values
    d_bw_poleidxs_newidxs = DistanceBWindexes(pole_indexes, indexes, d_EW_between, d_NS_between)
    min_dstanc = d_bw_poleidxs_newidxs.min()
    x_index, y_index = np.where(d_bw_poleidxs_newidxs == min_dstanc)
    base_ID = pole_ids[x_index[0]]
    conn_pole = pole_indexes[x_index[0], :]
    source = np.array([conn_pole[0]*d_EW_between, conn_pole[1]*d_NS_between])
    X = indexes*np.array([d_EW_between, d_NS_between])
    dists = np.sqrt(np.sum(np.power(X - source, 2), axis=1))
    dists_copy = np.copy(dists)
    dists_copy.sort()
    C = np.zeros((len(indexes), 2))
    ordered_index = [np.where(dists == val)[0][0] for val in dists_copy]
    C[:, 0:2] = indexes[ordered_index]
    indexes_IDs = []
    for i in range(len(C)):
        indexes_IDs.append(base_ID + 'S'+str(i+1))
    nw_indxs = np.vstack([C, conn_pole])
    nw_classes = PoleAngleClass(nw_indxs, d_EW_between, d_NS_between, type_)
    nw_classes = nw_classes[:-1]
    el, utmx, utmy, gps = PoleElevation(gen_LON, gen_LAT, gen_indexes, C, d_EW_between, d_NS_between)
    nw_classes['distance_from_source'] = dists_copy
    nw_classes['ID'] = indexes_IDs
    nw_classes['elevation'] = el
    nw_classes['UTM_X'] = utmx
    nw_classes['UTM_Y'] = utmy
    nw_classes['Type'] = type_
    nw_classes['GPS_X'] = gps[:, 0]
    nw_classes['GPS_Y'] = gps[:, 1]
    new_poleclasses = pd.concat([poleclasses, nw_classes])
    
    GDF = gpd.GeoDataFrame(new_poleclasses, 
                           geometry=gpd.points_from_xy(new_poleclasses.index_x*d_EW_between,
                                                       new_poleclasses.index_y*d_NS_between))
    
    from_idxs_net = networklength.filter(items=['index_x_from', 'index_y_from']).to_numpy()
    to_idxs_net = networklength.filter(items=['index_x_to', 'index_y_to']).to_numpy()
    from_coords = from_idxs_net*np.array([d_EW_between, d_NS_between])
    to_coords = to_idxs_net*np.array([d_EW_between, d_NS_between])
    Lines = [LineString((from_coords[i, :], to_coords[i, :])) for i in range(len(to_coords))]
    if type_ == 'MV':
        net_len = NetworkLength(nw_indxs, d_EW_between, d_NS_between)
        net_len['Type'] = 'MV'
    else:
        ntlength = networklength[networklength.Type != 'MV']
        id_ = None
        for typ in ntlength.Type.unique():
            nt = ntlength[ntlength.Type == typ]
            
            from_ = np.zeros((len(nt), 2))
            from_[:, 0] = nt.index_x_from.values
            from_[:, 1] = nt.index_y_from.values
            
            to_ = np.zeros((len(nt), 2))
            to_[:, 0] = nt.index_x_to.values
            to_[:, 1] = nt.index_y_to.values
            
            from_minus = np.abs(from_ - conn_pole)
            to_minus = np.abs(to_ - conn_pole)
            
            if from_minus.min(axis=0)[0] == 0 and from_minus.min(axis=0)[1] == 0:
                id_ = typ
                break
            if to_minus.min(axis=0)[0] == 0 and to_minus.min(axis=0)[1] == 0:
                id_ = typ
                break
        net_len = NetworkLength(nw_indxs, d_EW_between, d_NS_between)
        net_len['Type'] = id_
        
     # Put nework lines into a dataframe and geodataframe
    from_indexes = np.zeros((len(net_len), 2))
    from_indexes[:, 0] = net_len.index_x_from.values
    from_indexes[:, 1] = net_len.index_y_from.values
    
    to_indexes = np.zeros((len(net_len), 2))
    to_indexes[:, 0] = net_len.index_x_to.values
    to_indexes[:, 1] = net_len.index_y_to.values
    el1, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_indexes, from_indexes, d_EW_between, d_NS_between)
    el2, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_indexes, to_indexes, d_EW_between, d_NS_between)
    pole_elevation = CorrectLengthAngle(net_len.length.values, el1, el2)
    net_len['elevation_angle'] = pole_elevation
    net_len['adj_length'] = net_len['length']/np.cos(np.radians(net_len.elevation_angle))
    
    networklength = gpd.GeoDataFrame(networklength, geometry=Lines)
    net_len = pd.concat([networklength, net_len])
    
    networkcost = NetworkCost(costs, poleclasses, networklines, droplines)
    
    drop_fro = droplines.filter(items=['index_x_from', 'index_y_from']).to_numpy()
    drop_to = droplines.filter(items=['index_x_to', 'index_y_to']).to_numpy()
    
    drop_fro = drop_fro*np.array([d_EW_between, d_NS_between])
    drop_to = drop_to*np.array([d_EW_between, d_NS_between])
    
    lines = [LineString((drop_fro[i, :], drop_to[i, :])) for i in range(len(drop_to))]
    droplines = gpd.GeoDataFrame(droplines, geometry=lines)
    
    indexes_conn = conns.to_numpy()
    T_associated_conns = indexes_conn*np.array([d_EW_between, d_NS_between])
    connections = pd.DataFrame({'index_x':indexes_conn[:, 0], 'index_y':indexes_conn[:, 1]})
    geometry = [Point(p[0], p[1]) for p in T_associated_conns]
    connections = gpd.GeoDataFrame(connections, geometry=geometry)
    
    ConcessionDetails(GDF, net_len, droplines, networkcost, connections, concession, conc_id)

#==============================================================================

#==============================================================================
if __name__ == "__main__":
    
    t0 = time.time()
    # Specify concession
    concession = 'Makebe'
    
    #Set Inputs for optimizations
    Net_Parameters = pd.read_excel('uGrid_Input.xlsx', sheet_name = 'Net')
    reformatScaler = int(Net_Parameters['reformatScaler'][0]) #parameter to decrease the resolution of image (speeds up processing)
    exclusionBuffer = int(Net_Parameters['exclusionBuffer'][0])#meters that poles need to be form exclusions (other poles, exclusions, and connections)
    MaxDistancePoleConn = int(Net_Parameters['MaxDistancePoleConn'][0])#(m) the maximum distance allowed for a pole to be from a connection
    minPoles = int(Net_Parameters['minPoles'][0])
    maxPoles = int(Net_Parameters['maxPoles'][0])
    range_limit = int(Net_Parameters['range_limit'][0])
    Cost_kWh =  Net_Parameters['Cost_kWh'][0]
    
    #Load Lat and Long Gen
    lat_Generation = m.radians(Net_Parameters['lat_Generation'][0])
    long_Generation = m.radians(Net_Parameters['long_Generation'][0])
    
    repeats_MV_clusters = 1
    repeats_LV_clusters = 1
    
    repeats_LV_poles = 1
    repeats_MV_poles = 1
    
    repeats_improved_solution = 3
    total_repeats_lookback = repeats_improved_solution*repeats_MV_poles*repeats_LV_poles
    
    Best_Total_Cost = 99999999999999999999 #large dummy number
    
    #Community Load Data
    LoadKW_MAK = pd.read_excel('LoadKW_MAK.xlsx',index_col=None, header=None)
    global kW_max
    kW_max = max(LoadKW_MAK[0])
    
    #Line Data
    LV = 220 #V
    MV = 6300 #V
    LV_kW = (220*130)/1000 #kW
    MV_kW = 1.2*1000 #kW
    LV_safetyfactor = 0.5
    LV_kW_safety = LV_kW*LV_safetyfactor
    
    #Load all Econ input
    Econ_Parameters = pd.read_excel('uGrid_Input.xlsx', sheet_name = 'Econ')
   
    probability_MV = int(Net_Parameters['prob_MV'][0])
    restoration_time_MV = int(Net_Parameters['restoration_time_MV'][0])

    
    probability_LV = int(Net_Parameters['prob_LV'][0])
    restoration_time_LV = int(Net_Parameters['restoration_time_LV'][0])
    
    print("About to collect village data")
    #Collect Village Data (connections and exclusion zones)
    indexes_conn, indexes_excl, height, width, d_EW_between, d_NS_between,Long_exc_max, Long_exc_min,Lat_exc_max, Lat_exc_min = CollectVillageData()
    num_conns = len(indexes_conn[:,0])
    load_per_conn = kW_max/num_conns
    
    Best_Total_Cost_List = []
    Total_Cost_List = []
    Total_Reliability_Cost = []
    NoDecrease = 0
    
    # Load Components/Assembly costs
    ComponentsCost = pd.read_excel('uGrid_Input.xlsx', sheet_name = 'NetComponentsCost')
    costs = ComponentsCost.UnitPrice.values
    wiring_cost_MV = costs[10]
    wiring_cost_LV = costs[11]
    
    #Cycle through for best solution
    for i in range(minPoles, maxPoles):
        for nrep_LV in range(repeats_LV_poles):
            #Clustering for LV Poles
            ConnPoles_LV, indexes_poles_LV = PolePlacement(reformatScaler,i,exclusionBuffer,range_limit,indexes_conn,indexes_excl,height,width,d_EW_between,d_NS_between)                    
            LV_pole_num = len(indexes_poles_LV[:,0])
            #Calculate the minimum number of MV Poles
            min_MV_poles = int(kW_max/LV_kW)
            nrep_MV = 0
            while nrep_MV < repeats_MV_poles and NoDecrease == 0:
                print("Number of Poles: "+str(i)+",and LV repeat is "+str(nrep_LV+1)+" out of "+str(repeats_LV_poles)+", MV repeat is "+str(nrep_MV+1)+" out of "+str(repeats_MV_poles)+".") 
                t_temp = time.time()
                #Clustering for MW Poles
                load_behind = 99999999999999
                MV_pole_num = np.copy(min_MV_poles)
                while load_behind > LV_kW_safety:
                    MV_pole_num += 1 # add MV pole until load behind is below limits
                    for repeat_num in range(repeats_MV_clusters):
                        ConnPoles_MV, indexes_poles_MV = PolePlacement(reformatScaler,MV_pole_num,exclusionBuffer,range_limit,indexes_poles_LV,indexes_excl,height,width,d_EW_between,d_NS_between)  
                        #Calculate load behind by clusters
                        load_behind, ConnPoles_LVMV = loadBehindPoles(ConnPoles_MV, ConnPoles_LV, num_conns,MV_pole_num,load_per_conn)
                        if load_behind < LV_kW_safety:
                            break #Load behind the poles is below the limit, continue to wiring layout
                

                #Calculate HV Wiring Layout
                #Find POI of generation
                closest_pole, indexes_gen, connection_point = POI_Pole(lat_Generation,long_Generation,Long_exc_min,Lat_exc_min,d_EW_between,d_NS_between,indexes_poles_MV, 50)
                
                #add POI_MV to wiring for MV
                indexes_Poles_MV_wPOI = np.concatenate((indexes_poles_MV, [indexes_gen]), axis=0)
                total_distance_MV, OnOff_MV, DistancesBWPoles_MV, num_conn_per_pole_MV, reliability_cost_MV, wire_cost_MV, total_time_MV = WiringAlg(ConnPoles_LVMV,probability_MV,Cost_kWh,restoration_time_MV,indexes_Poles_MV_wPOI,wiring_cost_MV,d_EW_between,d_NS_between,"False")
                
                #Calculate LV Wiring Layout
                #Group indexes according to connecting MV pole
                group_indexes = []
                group_poles = []
                for MV_pole in range(MV_pole_num):
                    temp_group_indexes = np.array([])
                    temp_group_poles = []
                    for LV_pole in range(LV_pole_num):
                        if ConnPoles_MV[LV_pole,0] == MV_pole:
                            if not list(temp_group_indexes):
                                temp_group_indexes = indexes_poles_LV[LV_pole,:]
                            else:
                                temp_group_indexes = np.vstack((temp_group_indexes, indexes_poles_LV[LV_pole,:]))
                            temp_group_poles.append(LV_pole)
                    temp_group_indexes = np.vstack((temp_group_indexes, indexes_poles_MV[MV_pole,:])) #add MV pole location to end
                    group_indexes.append(temp_group_indexes)
                    group_poles.append(temp_group_poles)
                
                
                #Determine wiring layout
                OnOff_groups = []
                DistanceBWPoles_groups = []
                total_distance_LV = 0
                reliability_cost_LV = []
                for MV_pole_group in range(MV_pole_num):
                    total_distance_temp, OnOff_temp, DistancesBWPoles_temp, num_conn_per_pole_temp, reliability_cost_temp, wire_cost_temp, total_time_temp = WiringAlg(ConnPoles_LV[:,0],probability_LV,Cost_kWh,restoration_time_LV,np.array(group_indexes[MV_pole_group]),wiring_cost_LV,d_EW_between,d_NS_between,"False")
                    #Add group OnOff to total OnOff
                    OnOff_groups.append(OnOff_temp)
                    DistanceBWPoles_groups.append(DistancesBWPoles_temp)
                    total_distance_LV = total_distance_LV + total_distance_temp
                    reliability_cost_LV.append(reliability_cost_temp)
                
                poleclasses, networklines, droplines, connections = ClassifyNetworkPoles(concession, lat_Generation, long_Generation,
                                    indexes_gen, indexes_Poles_MV_wPOI, group_indexes,
                                    OnOff_MV, DistancesBWPoles_MV, indexes_excl,
                                    70, range_limit, width, height, d_EW_between,
                                    d_NS_between, indexes_conn)

                BestNetworkCost = NetworkCost(costs, poleclasses, networklines, droplines)
                ConcessionDetails(poleclasses, networklines, droplines, BestNetworkCost, connections,
                                  concession)

    #             total_reliability_cost_LVMV = sum(reliability_cost_LV) + reliability_cost_MV
    #
    #             #print("Terminate")
    #             #Calculate Costs
    #             Total_Cost = PenaltiesToCost(total_reliability_cost_LVMV, costs, poleclasses, networklines, droplines)
    #             #Save Best Solution Based on Cost
    #             if Total_Cost < Best_Total_Cost and max(ConnPoles_LV[:,1]) < MaxDistancePoleConn:
    #                 #Save Best Solutions
    #                 print("Found best solution")
    #                 BestPoleClasses = poleclasses
    #                 BestNetworkLines = networklines
    #                 BestDropLines = droplines
    #                 Best_Total_Cost = Total_Cost
    #             print("This iteration's total cost is $"+str(int(Total_Cost))+" and the best total cost is $"+str(Best_Total_Cost)+".")
    #             print("This iteration's maximum distance between a house and a LV pole is "+str(round(max(ConnPoles_LV[:,1]),2))+" and the allowable distance is "+str(MaxDistancePoleConn)+".")
    #             temp_time = time.time() - t_temp
    #             print("This iteration's calculation time is "+str(round(temp_time,2))+"s.")
    #             print("************************************************************************")
    #
    #             #If solution hasn't improved, stop iterations
    #             Total_Reliability_Cost.append(total_reliability_cost_LVMV)
    #             Total_Cost_List.append(Total_Cost)
    #             total_iterations = len(Total_Cost_List)
    #             if total_iterations > total_repeats_lookback:
    #                 if min(Total_Cost_List[total_iterations-total_repeats_lookback:total_iterations]) > Best_Total_Cost:
    #                     #If solution hasn't improved in the last total repeats lookback count exit iterations
    #                     NoDecrease = 1
    #
    #             #add iteration to while loop
    #             nrep_MV += 1
    #
    # BestNetworkCost = NetworkCost(costs, BestPoleClasses, BestNetworkLines, BestDropLines)
    # ConcessionDetails(BestPoleClasses, BestNetworkLines, BestDropLines, BestNetworkCost, connections, concession)
    #
    # t1 = time.time()
    # total_time = t1-t0
    # print("The total calculation time is "+str(round(total_time,2))+".")