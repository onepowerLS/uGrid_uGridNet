#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Oct 23 02:36:16 2021

@author: Phylicio, Lefu
"""
import glob
import os
import warnings
import sys
import numpy as np
import pandas as pd
import geopandas as gpd
import math as m
import networkx as nx
import datetime as dt
import matplotlib.pyplot as plt
import k_means_constrained as kmc
from scipy.spatial import distance_matrix
from pdf2image import convert_from_path
from PIL import Image
from shapely.geometry import Point, LineString
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as img

site_name = concession = sys.argv[1]
concession_id = sys.argv[2]

def get_8760(village_name):
    filtered_list = glob.glob(f'{village_name}*8760*.xlsx')
    for f in filtered_list:
        if village_name in f and '8760' in f:
            return f
    return None

#==============================================================================
# Evaluate intersection points of 2D arrays
def Intersect2d(X1, X2):
    X_t1, X_t2 = [], []
    for x in X1:
        X_t1.append(tuple(x))
    for y in X2:
        X_t2.append(tuple(y))
    return np.array(list(set(X_t1).intersection(set(X_t2))))
#==============================================================================


#==============================================================================
#Calculate Distance between GPS coordinates with Haversine Formula (returns in m)
def GPStoDistance(Lat1,Lat2,Long1,Long2):
    R_earth = 6371 #earth's mean radius in km
    a = m.sin((Lat1-Lat2)/2)**2 + m.cos(Lat1)*m.cos(Lat2)*m.sin((Long1-Long2)/2)**2
    c = 2*m.atan2(m.sqrt(a),m.sqrt(1-a))
    d = (R_earth*c)*1000 #m
    return d
#===============================================================================


#==============================================================================
# Exclusion Mapper using 2nd array instead of increasing list of indexes
def ExclusionMapper(ExclusionMap_array,reformatScaler,exclusionBuffer,d_EW_between,d_NS_between,width_new,height_new):
# This function takes the exclusion array, collected from the pdf image, and 
# recasts the exclusion zones into a new array accounting the for the exclusion 
# buffer and reformat scaler. The reformat scaler is used to reduce the 
# dimensions of the exclusion array so the program can run faster. 
# Additionally, it can make sense to reduce the resolution if the original pixel 
# resolution is less than a foot, such a high resolution is unnecessary. 
    print("in exclusion mapper")
    #Extend the exclusions to include the buffer zone
    bufferArrayWidth_EW = m.ceil(exclusionBuffer/d_EW_between) #reformatted indexes array width 
    bufferArrayHeight_NS = m.ceil(exclusionBuffer/d_NS_between)
    height_og = len(ExclusionMap_array[:,0,0])
    width_og = len(ExclusionMap_array[0,:,0])
    new_exclusions = np.zeros((width_new,height_new))
    indexes = []
    #t0 = time.time()
    
    for i in range((width_og-1),-1,-1):
        k = int((width_og-1-i)/reformatScaler)
        for j in range(0,height_og,1):
            l = int(j/reformatScaler)
            if ExclusionMap_array[j,i,0] != 255: #exclusion zones are both grey and black, safe zones are white (255,255,255)
                #set everything within buffer to 1
                if k < bufferArrayWidth_EW:
                    if l < bufferArrayHeight_NS:
                        new_exclusions[0:k+bufferArrayWidth_EW,0:l+bufferArrayHeight_NS] = 1
                    elif l > height_og-bufferArrayHeight_NS:
                        new_exclusions[0:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:height_new] = 1
                    else:
                        new_exclusions[0:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:l+bufferArrayHeight_NS] = 1
                elif k > width_new-bufferArrayWidth_EW:
                    if l < bufferArrayHeight_NS:
                        new_exclusions[k-bufferArrayWidth_EW:width_new,0:l+bufferArrayHeight_NS] = 1
                    elif l > height_og-bufferArrayHeight_NS: 
                        new_exclusions[k-bufferArrayWidth_EW:width_new,l-bufferArrayHeight_NS:height_new] = 1
                    else:
                        new_exclusions[k-bufferArrayWidth_EW:width_new,l-bufferArrayHeight_NS:l+bufferArrayHeight_NS] = 1
                else:
                    if l < bufferArrayHeight_NS:
                        new_exclusions[k-bufferArrayWidth_EW:k+bufferArrayWidth_EW,0:l+bufferArrayHeight_NS] = 1
                    elif l > height_og-bufferArrayHeight_NS: 
                        new_exclusions[k-bufferArrayWidth_EW:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:height_new] = 1
                    else:
                        new_exclusions[k-bufferArrayWidth_EW:k+bufferArrayWidth_EW,l-bufferArrayHeight_NS:l+bufferArrayHeight_NS] = 1
    del ExclusionMap_array
    print("finished remapping")
    new_exclusions = np.flip(new_exclusions)
    
    for o in range(width_new):
        for p in range(height_new):
            if new_exclusions[o,p] == 1:
                indexes.append([o,p])
    indexes = np.array(indexes)
    print("finished new indexing")

    #Save new indexes
    index_csv_name = "indexes_reformatted_%s_bufferzone_%s.csv" %(str(reformatScaler),str(exclusionBuffer))
    np.savetxt(index_csv_name,indexes, delimiter=",")
    return indexes
#==============================================================================


#=============================================================================
# Get Distance between array indexes
def DistanceBWindexes(indexesA, indexesB, d_EW_between, d_NS_between):
# This function calculates the distances between array indexes. This works for 
# any array, just the distance between indexes (or “pixels”) needs to be inputted.
    if type(indexesA) == list:
        indexesA_ = np.array(indexesA)*np.array([d_EW_between, d_NS_between]) 
    else:
        indexesA_ = indexesA*np.array([d_EW_between, d_NS_between]) 
    if type(indexesB) == list:
        indexesB_ = np.array(indexesB)*np.array([d_EW_between, d_NS_between]) 
    else:
        indexesB_ = indexesB*np.array([d_EW_between, d_NS_between])
    if len(indexesA_.shape) == 1:
        indexesA_ = [indexesA_.tolist()]
    else:
        indexesA_ = indexesA_.tolist()
    if len(indexesB_.shape) == 1:
        indexesB_ = [indexesB_.tolist()]
    else:
        indexesB_ = indexesB_.tolist()
    DistanceAB = distance_matrix(indexesA_, indexesB_)
    return DistanceAB
#=============================================================================

        
#==============================================================================
# Clustering of Connections for Initial Solution Pole Placement
def Clustering(indexes_conn, num_clusters):
# This function clusters the house location to obtain initial placements of 
# the LV (220V) poles. The number of clusters to create is set by the 
# num_clusters input. The clustering is performed with gaussian mean clustering 
# using the scikitlearn GaussianMixture function. 
    from sklearn import mixture
    
    X = np.copy(indexes_conn)

    cv_type = 'tied'
    gmm = mixture.GaussianMixture(n_components = num_clusters,covariance_type=cv_type)
    gmm.fit(X)
    clf = gmm 
    Y_ = clf.predict(X) #Y_ the index is the connection # and the value is the pole
    #gmm.means_: the index is the pole #, the numbers are the indexes of that pole
    
    #round and convert means to integers
    means = np.copy(gmm.means_) #the location (indexes) of the mean center of the cluster, where the pole is placed
    for i in range(len(gmm.means_)):
        means[i][0] = int(means[i][0])
        means[i][1] = int(means[i][1])
        
    clusters = []
    for cluster in range(num_clusters):
        clusters.append(X[np.where(Y_==cluster), :][0])
    return clusters, means
#==============================================================================


#==============================================================================
# Find non-exlcusion spot in growing circular manner
def FindNonExclusionSpot(index_x_og, index_y_og, index_excl_comp,range_limit,max_y,max_x):
# This function determines if the initial pole placement is in an exclusion zone. 
# If the pole is in an exclusion zone, the function tests the area around the 
# initial placement in a circular fashion to find the closest index location 
# to the initial placement that is not an exclusion zone.     
    for i in range(range_limit):
        #Define limits
        x_min = int(index_x_og - i)
        if x_min < 0:
            x_min = 0
        x_max = int(index_x_og + i)
        if x_max > max_x:
            x_max = max_x
        y_min = int(index_y_og - i)
        if y_min < 0:
            y_min = 0
        y_max = int(index_y_og + i)
        if y_max > max_y:
            y_max = 0
        #Testing around spot
        #xmin and xmax
        for y in range(y_min,y_max):
            index_pole_comp = x_min + (y*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x_min, y
            index_pole_comp = x_max + (y*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x_min, y
        #ymin and ymax
        for x in range(x_min,x_max):
            index_pole_comp = x + (y_min*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x, y_min
            index_pole_comp = x + (y_max*0.00001)
            if not index_pole_comp in index_excl_comp:
                return x, y_max
    
    #if nothing found in within range limit, return dummy number
    return 9000000,0
#==============================================================================


#==============================================================================
# Matching poles and connections by closest distance
def MatchPolesConn(indexes_conn,indexes_poles,d_EW_between,d_NS_between):
# This function creates an array to specify which poles connect to which houses, 
# by which houses are closest to which poles.    
    ConnPoles = np.zeros((len(indexes_conn[:,0]),2))      
    DistanceConnPoles = DistanceBWindexes(indexes_conn,indexes_poles,d_EW_between,d_NS_between)
    for i in range(len(ConnPoles[:,0])):
        ConnPoles[i,0] = np.argmin(DistanceConnPoles[i,:]) #The index of the closest pole for this connection, as indexed in "indexes_poles"
        ConnPoles[i,1] = np.min(DistanceConnPoles[i,:]) #The distance to the closest pole for this connection
    return ConnPoles
#==============================================================================


#==============================================================================
# k-means constrained clustering
def KMeansConstrained(X, n_clusters, min_cluster_size, max_cluster_size):
    labels = kmc.KMeansConstrained(n_clusters, min_cluster_size, max_cluster_size).fit_predict(X)
    df = pd.DataFrame({'x':X[:, 0], 'y':X[:, 1], 'labels':labels})
    clusters, means = [], np.zeros((n_clusters, 2))
    for idx, lab in enumerate(df.labels.unique()):
        Xtemp = df[df.labels == lab].filter(items=['x', 'y']).to_numpy()
        clusters.append(Xtemp)
        means[idx, :] = Xtemp.mean(axis=0)
    return clusters, means
#==============================================================================


#==============================================================================
# Cluster customers together given number of transformers
def ClusterTransformerCusts(indexes_conn, d_EW_between, d_NS_between, num_trans,
                            min_trans_custs, max_trans_custs):
    """
    Determine the clusters of connections using the k-means-constrained algorithm
    """
    conns = indexes_conn*np.array([d_EW_between, d_NS_between])
    clusters, centroids = KMeansConstrained(conns, num_trans, min_trans_custs,  max_trans_custs)
    transformed_clusters = []
    for cluster in clusters:
        c = cluster*np.array([1/d_EW_between, 1/d_NS_between])
        c = c.astype(int)
        transformed_clusters.append(c)
    
    transformed_centroids = centroids*np.array([1/d_EW_between, 1/d_NS_between])
    transformed_centroids = transformed_centroids.astype(int)
    return transformed_clusters, transformed_centroids
#==============================================================================


#==============================================================================
# Initial Placement of LV connection poles 
def LVPolesPlacement(customer_clusters, d_EW_between, d_NS_between, index_excl_comp, 
                     range_limit, max_x, max_y, max_d_BW_pole_conn):
    poles_loc = []
    for cluster in customer_clusters:
        clust = cluster*np.array([d_EW_between, d_NS_between])
        n_clusters = max(1, len(clust)//6 + 1) # Determine the intial pole number
        custs, poles = KMeansConstrained(clust, n_clusters, 2, 6)
        dists = distance_matrix(clust, poles)
        max_d = dists.min(axis=1).max()
        while max_d > max_d_BW_pole_conn:
            #print("Long distance: {}, repeat".format(dists.min(axis=1).max()))
            n_clusters += 1
            try:
                custs, poles = KMeansConstrained(clust, n_clusters, 2, 6)
                dists = distance_matrix(clust, poles)
                if max_d == dists.min(axis=1).max(): # if the result does not improve, terminate
                    break
                else:
                    max_d = dists.min(axis=1).max()
            except:
                break
        poles = poles*np.array([1/d_EW_between, 1/d_NS_between])
        poles = poles.astype(int)
        for k, pt in enumerate(poles):
            x, y = FindNonExclusionSpot(pt[0], pt[1], index_excl_comp, range_limit, 
                                            max_y, max_x)
            poles[k, :] = np.array([x, y])
        poles_loc.append(poles)
    return poles_loc
#==============================================================================


#==============================================================================
# Evaluate Transformer Poles Locations on the MV Network
def TransformerLocations(LV_Poles_Clusters, d_EW_between, d_NS_between, index_excl_comp, 
                     range_limit, max_x, max_y):
    trans_loc = np.zeros((len(LV_Poles_Clusters), 2))
    for idx, cluster in enumerate(LV_Poles_Clusters):
        c = cluster*np.array([d_EW_between, d_NS_between])
        trans_loc[idx, :] = c.mean(axis=0)
    trans_loc = trans_loc*np.array([1/d_EW_between, 1/d_NS_between])
    trans_loc = trans_loc.astype(int)
    
    for k, pt in enumerate(trans_loc):
        x, y = FindNonExclusionSpot(pt[0], pt[1], index_excl_comp, range_limit, 
                                        max_y, max_x)
        trans_loc[k, :] = np.array([x, y])
    return trans_loc
#==============================================================================


#==============================================================================
# Collect the connection and exclusions data
def CollectVillageData(site_name, reformatScaler=1, exclusionBuffer=2, max_d = 4000):
    # This function collects all of the data of the community needed to determine the network layout.

    # Gather the information needed
    # Import csv file which has been converted from the klm file
    # This gives the points of connections which are houses to link to the distribution grid
    Connect_nodes = pd.read_excel(site_name + '_connections.xlsx')
    Exclusion_nodes = pd.read_excel(site_name + '_exclusions.xlsx', sheet_name='exclusions')
    
    # Plot Connections
    #fig, ax = plt.subplots(dpi=150)
    #ax.scatter(Connect_nodes.Longitude.values, Connect_nodes.Latitude.values)
    #plt.show()
    

    # Identify gps coordinate min and max to determine coordinates of edges of jpg image
    Longitude_exc = Exclusion_nodes['x']
    Latitude_exc = Exclusion_nodes['y']
    # also convert these degrees to radians
    Lat_exc_min = m.radians(Latitude_exc.min())  # top of image (north)
    Lat_exc_max = m.radians(Latitude_exc.max())  # bottom of image (south)
    Long_exc_min = m.radians(Longitude_exc.min())  # left of image (east)
    Long_exc_max = m.radians(Longitude_exc.max())  # right of image (west)

    # Calculate the distance between the gps coordiantes using Haversine Formula
    # North South Distance #measuring latitude difference
    d_NS = GPStoDistance(Lat_exc_max, Lat_exc_min, Long_exc_max, Long_exc_max)  # m
    # East West Distance #measuring longitude difference
    d_EW = GPStoDistance(Lat_exc_max, Lat_exc_max, Long_exc_max, Long_exc_min)  # m
    
    if d_NS > max_d or d_EW > max_d:
        warnings.warn("Warning! The distances seem too high, you may want to check"\
                      +" your input coordinates. Code likely to take long to execute"\
                      +" the loops")
    # Load Files
    #load_file = get_8760(site_name)
    #Load = pd.read_excel(load_file, sheet_name='8760')

    #TODO: MSO defined Peakload after doing a data fit
    #PeakLoad = len(len(indexes_conn))*(0.8957*(len(indexes_conn))**(-0.243))
    PeakLoad = 5
    print("PeakLoad is {}".format(PeakLoad))

    # Import kml pdf file (of exclusions) and convert to jpg
    pages = convert_from_path(site_name + '_exclusions.pdf', 500)
    for page in pages:
        page.save(site_name+ '_exclusions.jpg', 'JPEG')

    # Convert JPG to array
    ExclusionMap = Image.open(site_name + '_exclusions.jpg')
    ExclusionMap_array = np.array(ExclusionMap)
    # Filter rgb value to 0 'non exclusion' and 1 'exclusion'
    # Black 0-0-0, White 255-255-255
    height = int(len(ExclusionMap_array[:, 0]) / reformatScaler)  # this is y_index_max
    width = int(len(ExclusionMap_array[0, :]) / reformatScaler)  # this is x_index_max
    filename = "index_maxes_%s.csv" % str(reformatScaler)
    np.savetxt(filename, [height, width], delimiter=",")

    # Determine distance between reformatted pixels (between values in the array)
    d_EW_between = d_EW / width  # m
    d_NS_between = d_NS / height  # m
    filename = "d_between_%s.csv" % str(reformatScaler)
    np.savetxt(filename, [d_EW_between, d_NS_between], delimiter=",")

    # Load exlusion map, if not available then perform
    # This gathers the exclusion array indexes
    try:
        # print("in try loop")
        index_csv_name = "indexes_reformatted_%s_bufferzone_%s.csv" % (str(reformatScaler), str(exclusionBuffer))
        indexes_excl = np.loadtxt(index_csv_name, delimiter=",")
    except:
        print("in except loop")
        # quit()
        indexes_excl = ExclusionMapper(ExclusionMap_array, reformatScaler, exclusionBuffer, d_EW_between, d_NS_between,
                                       width, height)

    # Match the connection locations to locations in the array
    # Find distance between east limit of image and connection
    try:
        index_csv_name = "indexes_conn_reformatted_%s.csv" % str(reformatScaler)
        indexes_conn = np.loadtxt(index_csv_name, delimiter=",")
    except:
        d_Econnection = np.zeros(len(Connect_nodes))
        d_Nconnection = np.zeros(len(Connect_nodes))
        indexes_conn = np.zeros((len(Connect_nodes), 2))
        for i in range(len(Connect_nodes)):  # iteration through connections
            d_Econnection[i] = GPStoDistance(Lat_exc_min, Lat_exc_min, Long_exc_min,
                                             m.radians(float(Connect_nodes['Longitude'][i])))  # m
            # distance of connection to the east (left) (x index)
            d_Nconnection[i] = GPStoDistance(Lat_exc_min, m.radians(float(Connect_nodes['Latitude'][i])), Long_exc_min,
                                             Long_exc_min)  # m
            # distance of connection to the north (top) (y index)
            # Get array index locations of all connections
            indexes_conn[i, 0] = int(d_Econnection[i] / d_EW_between)
            # print(indexes_conn[i,0])
            indexes_conn[i, 1] = int(d_Nconnection[i] / d_NS_between)
            # print(indexes_conn[i,1])
            index_csv_name = "indexes_conn_reformatted_%s.csv" % str(reformatScaler)
            np.savetxt(index_csv_name, indexes_conn, delimiter=",")
    #MSO did a data fit to determine the equation below. 
    load_per_conn = 0.8957*(len(indexes_conn))**(-0.243)
    print("Load per connection is {}".format(load_per_conn))
    
    # Network Inputs
    net_inputs = pd.read_excel(site_name + '_uGrid_Input.xlsx', sheet_name='Net')
    
    # Financial inputs
    financial_inputs = pd.read_excel(site_name + '_uGrid_Input.xlsx', sheet_name='NetComponentsCost')
    financial_params = dict(zip(financial_inputs.Component.values, financial_inputs.UnitPrice.values))
    costs = financial_inputs.UnitPrice.values
    
    # Voltage Drop Inputs
    vdrop_inputs = pd.read_excel(site_name + '_uGrid_Input.xlsx', sheet_name='VoltageDrop')
    
    results = {'indexes_conn':indexes_conn, 'indexes_excl':indexes_excl,
               'height':height, 'width':width, 'd_EW_between':d_EW_between,
               'd_NS_between':d_NS_between, 'Long_exc_max':Long_exc_max,
               'Long_exc_min':Long_exc_min, 'Lat_exc_max':Lat_exc_max,
               'Lat_exc_min':Lat_exc_min, 'load_per_conn':load_per_conn,
               'peak_demand':PeakLoad,
               'reformatScaler':int(net_inputs['reformatScaler'][0]),
               'exclusionBuffer':int(net_inputs['exclusionBuffer'][0]),
               'MaxDistancePoleConn':int(net_inputs['MaxDistancePoleConn'][0]),
               'MaxDistancePoleLV':int(net_inputs['MaxDistancePoleLV'][0]),
               'MaxDistancePoleMV':int(net_inputs['MaxDistancePoleMV'][0]),
               'range_limit':int(net_inputs['range_limit'][0]),
               'lat_Generation':m.radians(net_inputs['lat_Generation'][0]),
               'long_Generation':m.radians(net_inputs['long_Generation'][0]),
               'derate':float(vdrop_inputs['Derate'][0]),
               'tl_power':int(vdrop_inputs['TI Power'][0]),
               'other_power_factor':float(vdrop_inputs['Other Power Factor'][0]),
               'mv_3ph_voltage':int(vdrop_inputs['MV 3-Phase Voltage'][0]),
               'lv_3ph_voltage':int(vdrop_inputs['LV 3-Phase Voltage'][0]),
               'household_voltage':int(vdrop_inputs['Household Voltage'][0]),
               'transformer_voltage':int(vdrop_inputs['Transformer Voltage'][0]),
               'max_drop_volt':float(vdrop_inputs['Max Voltage Drop'][0]),
               'min_household_voltage':int(vdrop_inputs['Household Voltage - Min'][0]),
               'breaker_rating_factor':float(vdrop_inputs['Breaker Rating Factor'][0]),
               'power_factor':float(vdrop_inputs['Power Factor'][0]),
               'constant':float(vdrop_inputs['Constant'][0]),
               'characteristic_length_factor':float(vdrop_inputs['Characteristic Length Factor'][0]),
               'costs':costs
               }
    
    results.update(financial_params)
    return results
#==============================================================================


#==============================================================================
# Calculate Closest Pole to POI (point of interconnection) to generation
def POI_Pole(lat_Generation,long_Generation,Long_exc_min,Lat_exc_min,d_EW_between,d_NS_between,indexes_poles,d_BW_Adj_Poles):
    EW_dis = GPStoDistance(Lat_exc_min,Lat_exc_min,Long_exc_min,long_Generation)
    NS_dis = GPStoDistance(Lat_exc_min,lat_Generation,Long_exc_min,Long_exc_min)
    EW_index = int(EW_dis/d_EW_between)
    NS_index = int(NS_dis/d_NS_between)
    indexes_gen = [EW_index,NS_index]
    #Calculate distances between generation and pole
    #Individually feed in each pair
    num_poles = len(indexes_poles[:,0])
    Distance_Gen_Poles = np.zeros(num_poles)    
    for i in range(num_poles):
        Distance = DistanceBWindexes(indexes_gen,indexes_poles[i,:],d_EW_between,d_NS_between) #Type error: only size-1 arrays can be converted to Python Scalars
        Distance_Gen_Poles[i] = Distance
    closest_pole = np.argmin(Distance_Gen_Poles)
    Distance_Gen_Poles[closest_pole] = 90000000000 # put dummy number
    closest_pole2 = np.argmin(Distance_Gen_Poles)
    edge_dist = d_BW_Adj_Poles/np.sqrt(d_EW_between**2 + d_NS_between**2)
    edge = [indexes_poles[closest_pole, :], indexes_poles[closest_pole2, :],
            np.sqrt(np.sum(np.power(indexes_poles[closest_pole2, :] - indexes_poles[closest_pole, :], 2)))]
    
    Points = IntermediatePoints(edge, edge_dist)
    new_distances = np.zeros(len(Points))
    for i in range(len(Points)):
        new_distances[i] =  DistanceBWindexes(indexes_gen, Points[i, :],d_EW_between,d_NS_between)
    idx = np.argmin(new_distances)
    return [closest_pole, closest_pole2], indexes_gen, Points[idx, :]
#==============================================================================


#==============================================================================
# Given an edge (p1, p2, d_p1_p2), and allowable distance between adjacent points
# create intermediate points along the straight line connecting the two nodes
def IntermediatePoints(edge, edge_dist, dropline=False):
    P1, P2, k = edge # unpack points
    if dropline == False:
        if k < 1.3*edge_dist:
            return np.vstack((np.array([P1[0], P1[1]]), np.array([P2[0], P2[1]])))

    nodes_num = int(k//edge_dist) # number of nodes to add
    
    if nodes_num == 0:
        nodes_num = 1
        
    h_incr = 0 # horizontal increment
    v_incr = 0 # vertical increment
    if P2[0] == P1[0]:
        if P2[1] > P1[1]:
            theta = np.pi/2
            v_incr = edge_dist
        else:
            theta = 3*np.pi/2
            v_incr = -edge_dist
    if P2[1] == P1[1]:
        if P2[0] > P1[0]:
            theta = 0
            h_incr = edge_dist
        else:
            theta = np.pi
            h_incr = -edge_dist
    if P2[0] != P1[0] and P2[1] != P1[1]:
        g = (P2[1] - P1[1])/(P2[0] - P1[0]) # evaluate gradient
        # Determine theta and direction
        theta = m.atan(g)
        if P2[0] > P1[0] and P2[1] > P1[1]: # 1st Quadrant 
            h_incr = edge_dist*m.cos(theta) 
            v_incr = edge_dist*m.sin(theta) 
            
        elif P2[0] < P1[0] and P2[1] > P1[1]: # 2nd Quadrant 
             h_incr = -edge_dist*m.cos(theta) 
             v_incr = -edge_dist*m.sin(theta) 
             
        elif P2[0] < P1[0] and P2[1] < P1[1]: # 3rd Quadrant 
             h_incr = -edge_dist*m.cos(theta) 
             v_incr = -edge_dist*m.sin(theta) 
             
        elif P2[0] > P1[0] and P2[1] < P1[1]: # 4th Quadrant 
             h_incr = edge_dist*m.cos(theta) 
             v_incr = edge_dist*m.sin(theta)
    
    new_nodes = np.array([P1[0], P1[1]])
    for i in range(nodes_num):
        x_i = P1[0] + (i + 1)*h_incr
        y_i = P1[1] + (i + 1)*v_incr
        node = np.array([int(x_i), int(y_i)])
        new_nodes = np.vstack((new_nodes, node))
    new_nodes = np.vstack((new_nodes, np.array([P2[0], P2[1]])))
    
    # If the distance between the last two poles is less than 30%, remove the 
    # last intermediate pole
    d_last = np.sqrt(np.sum(np.power(new_nodes[-1, :] - new_nodes[-2,:], 2)))
    if d_last < 0.15*edge_dist:
        new_nodes = np.vstack((new_nodes[:-3, :], new_nodes[-1, :]))
    return new_nodes
#==============================================================================



#==============================================================================
# Evaluate Intermediate Poles respecting exclusion zones
def IntermediatePoles(indexes, OnOff_, d_BW_Poles, d_BW_Adj_Poles, index_excl_comp, range_limit, max_y, max_x):
    
    OnOff = np.triu(OnOff_) # Take connectivity from the upper triangle (above diagonal)
    # Taking the lower triangle would reverse starting points and end points
    # and would have to be indexed by columns instead of rows
    pairs = []
    for i, index in enumerate(indexes):
        idxs = np.where(OnOff[i, :] == 1)[0]
        if not list(idxs):
            pass
        else:
            for idx in idxs:
                d = d_BW_Poles[i][idx]
                pairs.append([index, indexes[idx, :], d])
                
    # Convert absolute distance to pixel distance
    edge = pairs[0]
    edge_d = np.sqrt(np.sum((edge[0] - edge[1])**2))
    pixel_d = edge_d/edge[2] # pixel distance per meter
    threshold = d_BW_Adj_Poles*pixel_d # 
    
    new_pairs = np.array([])
    for e in pairs:
        e[2] = e[2]*pixel_d
        if not list(new_pairs):
            new_pairs = IntermediatePoints(e, threshold)
        else:
            new_pairs = np.vstack((new_pairs, IntermediatePoints(e, threshold)))
    new_pairs = np.unique(new_pairs, axis=0) # Remove repeating points
    
    # Respect the exclusion zones
    for i, pt in enumerate(new_pairs):
        if pt[1] in indexes[:, 1] and pt[0] in indexes[:, 0]:
            pass
        else:
            x, y = pt[0], pt[1]
            #x, y = FindNonExclusionSpot(pt[0], pt[1], index_excl_comp, range_limit, max_y, max_x)
            new_pairs[i, :] = np.array([x, y])
        
    return  new_pairs
#==============================================================================


#==============================================================================
# Create Intermediate LV connection poles for those with longer distances
def IntermediateConnectionLVPoles(customer_clusters, LVPoleClusters, d_EW_between,
                                  d_NS_between, index_excl_comp, range_limit, max_x,
                                  max_y, d_BW_Adj_Poles):
    extended_pole_clusters = []
    for idx_, customer_cluster in enumerate(customer_clusters):
        pole_cluster = LVPoleClusters[idx_]
        generated_points = pole_cluster.copy()
        dists = DistanceBWindexes(customer_cluster, pole_cluster, d_EW_between, d_NS_between)
        min_dists = dists.min(axis=1) # Find distances to customers
        idxs = np.where(min_dists > d_BW_Adj_Poles)[0] # index of far customers
        for idx in idxs: 
            ix, idy = np.where(dists == min_dists[idx])
            pole = pole_cluster[idy[0], :] # corresponding pole
            cust = customer_cluster[idx, :] # corresponding customer
            # convert from pixels to metres
            pole = pole*np.array([d_EW_between, d_NS_between])
            cust = cust*np.array([d_EW_between, d_NS_between])
            # Put this into an edge with distance
            edge = (pole, cust, min_dists[idx])
            new_points = IntermediatePoints(edge, d_BW_Adj_Poles, dropline=True)
            new_points = new_points[1:-1]
            if len(new_points) != 0:
                # convert back to pixels
                new_points = new_points*np.array([1/d_EW_between, 1/d_NS_between])
                new_points = new_points.astype(int) # 
                # Ensure they obey exclusion zones
                for i, pt in enumerate(new_points):
                    x, y = FindNonExclusionSpot(pt[0], pt[1], index_excl_comp, range_limit, max_y, max_x)
                    new_points[i, :] = np.array([x, y])
                generated_points = np.vstack([generated_points, new_points])
        extended_pole_clusters.append(generated_points)
    return extended_pole_clusters
#==============================================================================


#==============================================================================
# Evaluate angle formed by three points
def AngleBWPoints(P_END_1, P_Mid, P_END_2, d_EW_between, d_NS_between):
    
    def angle(P1, P2):
        if P2[0] == P1[0] and P2[1] > P1[1]:
            return np.pi/2
        elif P2[0] == P1[0] and P2[1] < P1[1]:
            return 3*np.pi/2
        elif P2[1] == P1[1] and P2[0] > P1[0]:
            return 0
        elif P2[1] == P1[1] and P2[0] < P1[0]:
            return np.pi
        else:
            g = (P1[1] - P2[1])*d_NS_between/((P1[0] - P2[0])*d_EW_between)
            theta = m.atan(g)
            
            if P2[1] > P1[1] and P2[0] > P1[0]: # First Quadrant
                return theta
            elif P2[1] > P1[1] and P2[0] < P1[0]: # Second Qudrant
                return np.pi - np.abs(theta)
            elif P2[1] < P1[1] and P2[0] < P1[0]: # Third Quadrant
                return theta + np.pi
            elif P2[1] < P1[1] and P2[0] > P1[0]: # Fourth Quadrant
                return 2*np.pi - np.abs(theta)
    theta1 = angle(P_Mid, P_END_1)
    theta2 = angle(P_Mid, P_END_2)
    angles = np.degrees([theta1, theta2])
    if abs(angles[1] - angles[0]) <= 180:
        return 180 - abs(angles[1] - angles[0])
    else: 
        return 180 - (360 - abs(angles[1] - angles[0]))
    
#==============================================================================

#==============================================================================
# Return Minimum Spanning Tree 
def MST(indexes, d_EW_between, d_NS_between):
    """
    This function implements a least cost network using Prim's minimum spanning 
    tree algorithm
    """
    points = list(zip(indexes[:, 0]*d_EW_between, indexes[:, 1]*d_NS_between))
    try:
        rand_point = np.random.choice(points)
    except:
        rand_point = points[0]
    fPts = [rand_point] # initialize arbitrary starting points
    pts = list(points.copy()) # create shallow copy of the points
    pts.remove(rand_point)
    cPts = [] # connected points
    while len(pts) != 0:
        dist = [] # distances to all points
        dmins = [] # minimum distances
        for pt in fPts:
            dist.append([((pt[0] - pti[0])**2 + (pt[1] - pti[1])**2)**(1/2) for pti in pts])
        for dlist in dist:
            dmins.append(min(dlist)) # add minimum distance to this list
        # determine the index of the list with minimum distance    
        list_idx = dmins.index(min(dmins)) 
        # find the index of the minimum distance in a list it is in
        mind_idx = dist[list_idx].index(min(dist[list_idx]))
        
        cPts.append([fPts[list_idx], pts[mind_idx]]) # add a pair of connected points
        fPts.append(pts[mind_idx])
        pts.remove(pts[mind_idx])
    G = nx.Graph()
    G.add_edges_from(cPts)
    return G
#==============================================================================


#==============================================================================
# Return Minimum Spanning Tree Connectivity matrix
def MSTConnectivityMatrix(indexes, d_EW_between, d_NS_between):
    points = list(zip(indexes[:, 0]*d_EW_between, indexes[:, 1]*d_NS_between))
    G = MST(indexes, d_EW_between, d_NS_between)
    connectivityMatrix = nx.adjacency_matrix(G, nodelist=points).todense()
    return connectivityMatrix
#==============================================================================


#==============================================================================
# Put poles into respective branches
def EvaluateBranches(source_index, indexes, d_EW_between, d_NS_between):
    source_node = (source_index[0]*d_EW_between, source_index[1]*d_NS_between) 
    G = MST(indexes, d_EW_between, d_NS_between)
    G.remove_node(source_node)
    branches = list(nx.connected_components(G))
    branches_ = []
    for b in branches:
        indexes_ = np.array(list(b))
        indexes_[:, 0] = indexes_[:, 0]/d_EW_between
        indexes_[:, 1] = indexes_[:, 1]/d_NS_between
        branches_.append(indexes_)
    return branches_
#==============================================================================


#==============================================================================
# Classify poles by angle
def PoleAngleClass(pole_indexes, d_EW_between, d_NS_between, ntype):
    connectivity = MSTConnectivityMatrix(pole_indexes, d_EW_between, d_NS_between)
    classes = []
    for i in range(len(pole_indexes)):
        neighbors = np.sum(connectivity[i, :])
        p_mid = pole_indexes[i, :]
        if neighbors == 1:
            classes.append('terminal')
        elif neighbors == 2:
            neighbors_idx = np.where(connectivity[i, :] == 1)[1]
            p_end_1 = pole_indexes[neighbors_idx[0], :]
            p_end_2 = pole_indexes[neighbors_idx[1], :]
            agl = AngleBWPoints(p_end_1, p_mid, p_end_2, d_EW_between, d_NS_between)
            if ntype == 'LV':
                if agl in range(0,2) or agl in range(178,182):
                    classes.append("mid_straight")
                elif agl < 45:
                    classes.append("mid_less_45")
                elif agl >= 45:
                    classes.append("mid_over_45")
            elif ntype == 'MV':
                if agl in range(0,2) or agl in range(178,182):
                    classes.append("mid_straight")
                elif agl < 30:
                    classes.append("mid_less_30")
                elif agl > 30:
                    classes.append("mid_over_30")
        elif neighbors > 2:
            neighbors_idx = np.where(connectivity[i, :] == 1)[1]
            p_ends = pole_indexes[neighbors_idx, :]
           
            d_BW_p_ends = DistanceBWindexes(p_ends, p_ends, d_EW_between, d_NS_between)
            d_BW_p_ends = d_BW_p_ends + np.diagflat(np.ones(neighbors)*99999999)
            # Evaluate for all combinations of middle pole and all neighbors
            angles = np.zeros(neighbors)
            for j in range(neighbors):
                n1 = p_ends[j, :]
                index_n2 = np.argmin(d_BW_p_ends[j, :])
                n2 = p_ends[index_n2, :]
                d_BW_p_ends[index_n2, j] = 999999
                angles[j] = AngleBWPoints(n1, p_mid, n2, d_EW_between, d_NS_between)
                
            agl = angles.min()
            if ntype == 'LV':
                if agl in range(0,2) or agl in range(178,182):
                    classes.append("mid_straight")
                elif agl < 45:
                    classes.append("mid_less_45")
                elif agl >= 45:
                    classes.append("mid_over_45")
            elif ntype == 'MV':
                if agl in range(0,2) or agl in range(178,182):
                    classes.append("mid_straight")
                elif agl < 30:
                    classes.append("mid_less_30")
                elif agl > 30:
                    classes.append("mid_over_30")
    classes_ = pd.DataFrame(pole_indexes, columns = ['index_x', 'index_y'])
    classes_['AngleClass'] = classes 
    return classes_
#==============================================================================


#==============================================================================
# Get elevation of pole indexes using the generation latitude and logitude as 
# reference point from the open access API open-elevation
def PoleElevation(gen_LON, gen_LAT, gen_indexes, target_indexes, d_EW_between, d_NS_between):
    import requests
    GPD = gpd.GeoDataFrame(geometry=[Point(np.degrees(gen_LON), np.degrees(gen_LAT))])
    GPD = GPD.set_crs("WGS84") # WGS84
    GPD = GPD.to_crs("EPSG:22289") # Transform to UTM (Mercator)
    X, Y = list(GPD.geometry.values[0].coords)[0]
    X_Shifts = (gen_indexes[0] - target_indexes[:, 0])*d_EW_between
    Y_Shifts = (gen_indexes[1] - target_indexes[:, 1])*d_NS_between
    DF = pd.DataFrame({'index_x':target_indexes[:,0], 'index_y':target_indexes[:,1],
                       'UTM_X':X + X_Shifts, 'UTM_Y':Y + Y_Shifts})
    new_GDF = gpd.GeoDataFrame(DF, geometry=gpd.points_from_xy(DF.UTM_X, DF.UTM_Y))
    new_GDF = new_GDF.set_crs("EPSG:22289")#EPSG:3857
    new_GDF = new_GDF.to_crs("WGS84") #EPSG:4326
    gps = np.array([[list(i.coords)[0][0], list(i.coords)[0][1]] for i in new_GDF.geometry.values])
    

    #print('\r\n*-\r\n Reference Coordinate: ', np.degrees(gen_LON),np.degrees(gen_LAT),'\r\n*')

    #Recalibrate the uGridnet coordinate outputs
    for idx in range(len(gps)):
        #lat           #long
        gps[idx][1] , gps[idx][0] = gpsRecalibration(np.degrees(gen_LAT),gps[idx][1],np.degrees(gen_LON),gps[idx][0])

    elevations = []
    url = "https://maps.googleapis.com/maps/api/elevation/json?locations="
    if len(new_GDF) <= 256: 
        for idx, pt in enumerate(new_GDF.geometry.values):
            lon, lat = list(pt.coords)[0]
            if idx < len(new_GDF) - 1:
                url = url + str(lat) + ',' + str(lon) + '|'
            else:
                url = url + str(lat) + ',' + str(lon)
        url = url + '&key=AIzaSyB7NCOraSbaIBDVg2-BU5D_mX_Q2BwZV2E'
        payload = {}
        headers = {}
        response = requests.request('GET', url, headers=headers, data=payload).json()
        elevations = [res['elevation'] for res in response['results']]
    else:
        print('Many locations:', len(gps))
        points = new_GDF.geometry.values
        range_ = int(len(new_GDF)//256 + 1)
        for i in range(range_):
            partial_pts = points[i*256:i*256+256]
            url = "https://maps.googleapis.com/maps/api/elevation/json?locations="
            for idx, pt in enumerate(partial_pts):
                lon, lat = list(pt.coords)[0]
                if idx < len(partial_pts) - 1:
                    url = url + str(lat) + ',' + str(lon) + '|'
                else:
                    url = url + str(lat) + ',' + str(lon)
            url = url + '&key=AIzaSyB7NCOraSbaIBDVg2-BU5D_mX_Q2BwZV2E'
            payload = {}
            headers = {}
            response = requests.request('GET', url, headers=headers, data=payload).json()
            elevations = elevations + [res['elevation'] for res in response['results']]
    return np.array(elevations),  list(DF.UTM_X), list(DF.UTM_Y), gps
#==============================================================================


#==============================================================================
# Evaluate the altitude angle
def CorrectLengthAngle(horizontald1_2, elevation1, elevation2):
    net_elevation = np.abs(elevation1 - elevation2)
    cosine = horizontald1_2/np.sqrt(net_elevation**2 + horizontald1_2**2)
    return [np.degrees(m.acos(cos_)) for cos_ in cosine]
#==============================================================================


#==============================================================================
# Evaluate network segments length on a 2D plane
def NetworkLength(indexes, d_EW_between, d_NS_between):
    mst = MST(indexes, d_EW_between, d_NS_between)
    edges =list(mst.edges)
    og_indexes = np.zeros((len(edges), 4))
    for i, e in enumerate(edges):
        X = np.array(e)
        og_indexes[i, 0] = X[0, 0]/d_EW_between
        og_indexes[i, 1] = X[0, 1]/d_NS_between
        og_indexes[i, 2] =  X[1, 0]/d_EW_between
        og_indexes[i, 3] = X[1, 1]/d_NS_between
    df = pd.DataFrame(data=og_indexes, columns=['index_x_from', 'index_y_from', 
                                                'index_x_to', 'index_y_to'])
    gdf = gpd.GeoDataFrame(df, geometry=[LineString(e) for e in edges])
    gdf['length'] = [s.length for s in gdf.geometry.values]
    return gdf
#==============================================================================


#==============================================================================
# Evaluate the cost of the network
def NetworkCost(costs, dfpoles, dfnet, dfdropline):
    result_df = pd.DataFrame()
    references = ['Assembly - Pole - MV - Start',
                  'Assembly - Step - Down - Transformer',
                  'Assembly - Pole - MV - Mid',
                  'Assembly - Pole - MV - Bend <30',
                  'Assembly - Pole - MV - Bend >30',
                  'Assembly - Pole - MV - End',
                  'Assembly - Pole - LV - Mid',
                  'Assembly - Pole - LV - Bend <60',
                  'Assembly - Pole - LV - Bend >60',
                  'Assembly - Pole - LV - End',
                  'Wire - MV',
                  'Wire - LV',
                  'Wire - LineDrop',
                  'Assembly - Meter']
   
    mv_ref = dfpoles[dfpoles.Type == 'MV']
    mv_ref = mv_ref[mv_ref.distance_from_source != 0]
    lv_ref = dfpoles[dfpoles.Type == 'LV']
    quantity = np.zeros(len(references))
    quantity[0] = 1
    num_transformers = 0
    for item in mv_ref.ID.values:
        if item[-1].isalpha() == True:
            num_transformers += 1
    quantity[1] = num_transformers
    
    mv_pole_counts = mv_ref.value_counts(subset=['AngleClass'])
    try:
        quantity[2] = mv_pole_counts['mid_straight']
    except:
        pass
    try:
        quantity[3] = mv_pole_counts['mid_less_30']
    except:
        pass
    try:
        quantity[4] = mv_pole_counts['mid_over_30']
    except:
        pass
    try:
        quantity[5] = mv_pole_counts['terminal']
    except:
        pass
    
    lv_pole_counts = lv_ref.value_counts(subset=['AngleClass'])
    try:
        quantity[6] = lv_pole_counts['mid_straight']
    except:
        pass
    try:
        quantity[7] = lv_pole_counts['mid_less_60']
    except:
        pass
    try:
        quantity[8] = lv_pole_counts['mid_over_60']
    except:
        pass
    try:
        quantity[9] = lv_pole_counts['terminal']
    except:
        pass
    
    # Lines length
    mvnet = dfnet[dfnet.Type == 'MV']
    lvnet = dfnet[dfnet.Type != 'MV']
    quantity[10] = mvnet.adj_length.values.sum()
    quantity[11] = lvnet.adj_length.values.sum()
    
    # Line drop to households
    quantity[12] = dfdropline.Linedrop.values.sum()
    quantity[13] = len(dfdropline)
    
    temp_df = pd.DataFrame({'Part Reference':references,
                            'Qty':quantity,
                            'Price (USD)':costs})
    
    result_df = temp_df[['Part Reference', 'Qty','Price (USD)']]
    result_df['Line Total (USD)'] = result_df['Qty']*result_df['Price (USD)']
    return result_df
#==============================================================================


#==============================================================================
# Evaluate the IDs of connected poles 
def LinesPoles(pole_classes, network_lines):
    pole_indexes = pole_classes.filter(items=['index_x', 'index_y']).to_numpy()
    pole_ids = pole_classes.ID.values
    line_poles = network_lines.filter(items=['index_x_from', 'index_y_from',
                                             'index_x_to', 'index_y_to']).to_numpy()
    from_IDs, to_IDs = [], []
    pidxs = pole_indexes[:, 0] + pole_indexes[:, 1] + pole_indexes[:, 0]*pole_indexes[:, 1]
    for vector_ in line_poles:
        fro_ = vector_[0:2][0] + vector_[0:2][1] + vector_[0:2][0]*vector_[0:2][1]
        to_ = vector_[2:][0] + vector_[2:][1] + vector_[2:][0]*vector_[2:][1] 
        fro_diff = np.abs(pidxs -fro_)
        to_diff = np.abs(pidxs - to_)
        from_IDs.append(pole_ids[np.argmin(fro_diff)])
        to_IDs.append(pole_ids[np.argmin(to_diff)])
    return from_IDs, to_IDs
#==============================================================================


#==============================================================================
# Evaluate the pole IDs of the poles from which drop lines come from
def MatchConnectionsPoleID(pole_classes, droplines, connections, d_EW_between, d_NS_between):
    # Xpc = pole_classes.filter(items=['index_x', 'index_y']).to_numpy()
    # Xdrop = droplines.filter(items=['index_x_from'])
    
    Xconn = connections.filter(items=['index_x', 'index_y']).to_numpy()
    Xconn = Xconn*np.array([d_EW_between, d_NS_between])
    Xconn_ = Xconn.sum(axis=1) + Xconn.prod(axis=1)
    
    Xdrop = droplines.filter(items=['index_x_from', 'index_y_from', 'index_x_to',
                                    'index_y_to']).to_numpy()
    Xdrop_from = Xdrop[:, 0:2]*np.array([d_EW_between, d_NS_between])
    Xdrop_to = Xdrop[:, 2:]*np.array([d_EW_between, d_NS_between])
    
    Xdropf = Xdrop_from.sum(axis=1) + Xdrop_from.prod(axis=1)
    Xdropt = Xdrop_to.sum(axis=1) + Xdrop_to.prod(axis=1)
    
    XPolesClasses = pole_classes[pole_classes.Type == 'LV']
    Xpclass = XPolesClasses.filter(items=['index_x', 'index_y']).to_numpy()
    Xpclass = Xpclass*np.array([d_EW_between, d_NS_between])
    Xpc = Xpclass.sum(axis=1) + Xpclass.prod(axis=1)
    
    poleids = XPolesClasses.ID.values
    cpids = []
    for idx, val in enumerate(Xconn_):
        ID = poleids[np.argmin(np.abs(Xpc - Xdropf[np.argmin(np.abs(Xdropt - val))]))]
        cpids.append(ID)
        
    dpids = []
    for idx, val in enumerate(Xdropf):
        ID = poleids[np.argmin(np.abs(Xpc - val))]
        dpids.append(ID)
    return cpids, dpids
#==============================================================================


#==============================================================================
#  Evaluate the network segment the lines belong to
def NetworkLinesID(networklines, base):
    linetypes = networklines.Type.values
    ID_from = networklines.Pole_ID_From.values
    ID_to = networklines.Pole_ID_To.values
    sub_net, branch = [], []
    for idx, type_ in enumerate(linetypes):
        if type_ == 'MV':
            sub_net.append('M')
            branch.append(1)
        else:
            sub_n = ID_from[idx].replace(base, '')[0]
            b = ID_from[idx].replace(base, '')[1]
            if sub_n == 'M':
                sub_n = ID_to[idx].replace(base, '')[0]
                b = ID_to[idx].replace(base, '')[1]
            sub_net.append(sub_n)
            branch.append(b)
    return sub_net, branch
#==============================================================================


#==============================================================================
# Classify the poles by concession, MV/LV, Angle 
def ClassifyNetworkPoles(concession, gen_LAT, gen_LON, gen_site_indexes, 
                         indexes_MV_Poles_wPOI, group_indexes_LV,
                         OnOff_MV, d_BW_Poles_MV, index_excl_comp, d_BW_Adj_Poles, 
                         range_limit, max_y, max_x, d_EW_between, d_NS_between,
                         indexes_conn, conc_id = None):
    base = concession.upper()[0:3]
    if conc_id != None:
        base = concession.upper()[0:3] + str(conc_id)
    # MV Pole  and LV Pole Naming
    MV_poles_names = []
    MV_Pole_indexes = np.array([])
    
    LV_Poles_names = []
    LV_Pole_indexes = np.array([])
    
    AngleClasses = pd.DataFrame()
    network_length = gpd.GeoDataFrame()
    droplines = pd.DataFrame()
    basen = base + 'M'
    gen_index = gen_site_indexes
    if not list(MV_Pole_indexes):
        MV_Pole_indexes = np.array([gen_index[0], gen_index[1], 0])
    else:
        MV_Pole_indexes = np.vstack((MV_Pole_indexes, np.array([gen_index[0], gen_index[1], 0])))
    MV_poles_names.append(basen + str(1))
    t_gen_idx = np.array([gen_index[0]*d_EW_between, gen_index[1]*d_NS_between])
    mv_poles = indexes_MV_Poles_wPOI
    Yt = mv_poles*np.array([d_EW_between, d_NS_between])
    ydists = np.sqrt(np.sum(np.power(Yt - t_gen_idx, 2), axis=1))
    ydists_copy = np.copy(ydists)
    ydists_copy.sort()
    A = np.zeros((len(mv_poles), 2))
    ordered_idxs = [np.where(ydists == val)[0][0] for val in ydists_copy]
    A[:, 0] = mv_poles[:, 0][ordered_idxs]
    A[:, 1] = mv_poles[:, 1][ordered_idxs]
    all_MV_Poles = IntermediatePoles(mv_poles, OnOff_MV, d_BW_Poles_MV, d_BW_Adj_Poles,
                                     index_excl_comp, range_limit, max_y, max_x)
    
    angle_class_ = PoleAngleClass(all_MV_Poles, d_EW_between, d_NS_between, 'MV')
    if len(AngleClasses) == 0:
        AngleClasses = angle_class_
    else:
        AngleClasses = pd.concat([AngleClasses, angle_class_])
    net_len = NetworkLength(all_MV_Poles, d_EW_between, d_NS_between)
    net_len['Type'] = 'MV'
    if len(network_length) == 0:
        network_length = net_len
    else:
        network_length = pd.concat([network_length, net_len])
    branches = EvaluateBranches(gen_index, all_MV_Poles, d_EW_between, d_NS_between)
    # Mark transformer poles (subnetworks start points)
    t_counter = 65 # why start at 65? letter A is represented by 65, B 66 and so on
    # Note the position of each subnetwork
    subnetworks = np.zeros((1, 3))
    #nonzero_index = []
    for j, branch in enumerate(branches):
        n = j*100
        Xt = branch*np.array([d_EW_between, d_NS_between])
        dists = np.sqrt(np.sum(np.power(Xt - t_gen_idx, 2), axis=1))
        dists_copy = np.copy(dists)
        dists_copy.sort()
        B = np.zeros((len(branch), 3))
        ordered_index = [np.where(dists == val)[0][0] for val in dists_copy]
        B[:, 0:2] = branch[ordered_index]
        B[:, 2] = dists_copy
        MV_Pole_indexes = np.vstack((MV_Pole_indexes, B))
        names_ = [basen + str(n + i) for i in range(2, len(branch)+2)]
        intersection_points = Intersect2d(np.round(A), np.round(branch))
        for k in range(len(intersection_points)):
            pos_row, pos_col = np.where(np.round(B[:, 0:2]) == intersection_points[k, :])
            names_[pos_row[0]] = names_[pos_row[0]] + str(chr(t_counter))
            if np.array_equal(subnetworks, np.array([[0, 0, 0]])) == True:
                subnetworks[0, :] = np.array([intersection_points[k, 0], 
                                          intersection_points[k, 1], 
                                          t_counter])
            else:
                subnetworks = np.vstack([subnetworks,
                                         np.array([intersection_points[k, 0], 
                                          intersection_points[k, 1], 
                                          t_counter])])
            t_counter += 1
        MV_poles_names += names_
    
    # Low Voltage Poles
    group = group_indexes_LV
    sub_network_counter = 65
    for id_, g in enumerate(group):
        net_len = NetworkLength(g, d_EW_between, d_NS_between)
        net_len['Type'] = 'LV' + str(id_ + 1)
        if len(network_length) == 0:
            network_length = net_len
        else:
            network_length = pd.concat([network_length, net_len])
        branch_counter = 65
        for a in np.round(A):
            del_idx_x, del_idx_y = np.where(np.round(g) == a)
            if len(del_idx_x) == 2 and del_idx_x[0] == del_idx_x[1]:
                # Find the subnetwork ID
                sub_idx, sub_idy = np.where(subnetworks[:, 0:2] == a)
                if not list(sub_idx):
                    pass
                else:
                    sub_network = int(subnetworks[sub_idx[0], 2])
                    temp_pole_class = PoleAngleClass(np.round(g), d_EW_between, d_NS_between, 'LV')
                    #temp_pole_class.remove(temp_pole_class[del_idx_x[0]])
                    if len(AngleClasses) == 0:
                        AngleClasses = temp_pole_class
                    else:
                        AngleClasses = pd.concat([AngleClasses, temp_pole_class])
                    branches_ = EvaluateBranches(a, np.round(g), d_EW_between, d_NS_between)
                    for b_ in branches_:
                        source = np.array([a[0]*d_EW_between, a[1]*d_NS_between])
                        X = b_*np.array([d_EW_between, d_NS_between])
                        dists = np.sqrt(np.sum(np.power(X - source, 2), axis=1))
                        dists_copy = np.copy(dists)
                        dists_copy.sort()
                        C = np.zeros((len(b_), 3))
                        ordered_index = [np.where(dists == val)[0][0] for val in dists_copy]
                        C[:, 0:2] = b_[ordered_index]
                        C[:, 2] = dists_copy
                        if not list(LV_Pole_indexes):
                            LV_Pole_indexes = C
                        else:
                            LV_Pole_indexes = np.vstack((LV_Pole_indexes, C))
                        LV_Poles_names += [base + chr(sub_network) + chr(branch_counter) + str(m+1) for m in range(len(b_))]
                        branch_counter += 1
        sub_network_counter += 1
        
    # Determine customer connections
    lv_poles = np.concatenate(group)
    del_idx_x, del_idx_y = [], []
    for ir in range(len(mv_poles)):
        x = mv_poles[ir, 0]
        y = mv_poles[ir, 1]
        idx_x = np.where(lv_poles[:, 0] == x)[0]
        idx_y = np.where(lv_poles[:, 1] == y)[0]
        if not list(idx_x) or not list(idx_y):
            pass
        elif idx_x[0] == idx_y[0]:
            del_idx_x.append(idx_x[0])
            del_idx_y.append(idx_y[0])
    lv_poles_ = np.delete(lv_poles, del_idx_x, axis=0)
    
    dist_poles_conns = DistanceBWindexes(lv_poles_, indexes_conn, d_EW_between, d_NS_between)
    min_dists = dist_poles_conns.min(axis=0)
    T_lv_poles = lv_poles_*np.array([d_EW_between, d_NS_between])
    T_associated_conns = indexes_conn*np.array([d_EW_between, d_NS_between])
    temp_from_idx, temp_to_idx, temp_from_idy, temp_to_idy, lines = [],[],[],[],[]
    
    for idy in range(len(indexes_conn)):
        idx = np.argmin(dist_poles_conns[:, idy])
        temp_from_idx.append(lv_poles_[idx, 0])
        temp_to_idx.append(indexes_conn[idy, 0])
        temp_from_idy.append(lv_poles_[idx, 1])
        temp_to_idy.append(indexes_conn[idy, 1])
        lines.append(LineString((T_lv_poles[idx, :], T_associated_conns[idy, :])))
    
    temp_df = pd.DataFrame({'index_x_from':temp_from_idx,
                            'index_y_from':temp_from_idy,
                            'index_x_to':temp_to_idx,
                            'index_y_to':temp_to_idy,
                            'Linedrop':min_dists})
    droplines = gpd.GeoDataFrame(temp_df, geometry=lines)
    
    # Put pole classes into a dataframe and geodataframe
    print(MV_Pole_indexes.shape)
    print(LV_Pole_indexes.shape)
    DF = pd.DataFrame(data=np.vstack((MV_Pole_indexes, LV_Pole_indexes)),
                      columns=['index_x', 'index_y', 'distance_from_source'])
    DF['ID'] = MV_poles_names + LV_Poles_names
    DF['Type'] = ['MV']*len(MV_poles_names) + ['LV']*len(LV_Poles_names)
    AngleClasses = AngleClasses.drop_duplicates(subset=['index_x', 'index_y'])
    
    DF = DF.sort_values(by=['index_x', 'index_y'])
    AngleClasses = AngleClasses.sort_values(by=['index_x', 'index_y'])
    DF['AngleClass'] = AngleClasses['AngleClass'].values
    
    gen_index = gen_site_indexes
    data = DF#[DF.concession == conc]
    target_indexes = data.filter(items=['index_x', 'index_y']).to_numpy()
    el, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_site_indexes, 
                                          target_indexes, d_EW_between, d_NS_between)
    DF['elevation'] = el
    DF['UTM_X'] = utm_x
    DF['UTM_Y'] = utm_y
    DF['GPS_X'] = list(gps[:, 0])
    DF['GPS_Y'] = list(gps[:, 1])
    poleclasses = gpd.GeoDataFrame(DF, geometry=gpd.points_from_xy(DF.index_x*d_EW_between,
                                                           DF.index_y*d_NS_between))
    
    # Put nework lines into a dataframe and geodataframe
    from_indexes = np.zeros((len(network_length), 2))
    from_indexes[:, 0] = network_length.index_x_from.values
    from_indexes[:, 1] = network_length.index_y_from.values
    
    to_indexes = np.zeros((len(network_length), 2))
    to_indexes[:, 0] = network_length.index_x_to.values
    to_indexes[:, 1] = network_length.index_y_to.values
    el1, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_site_indexes, 
                                           from_indexes, d_EW_between, d_NS_between)
    el2, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_site_indexes, 
                                           to_indexes, d_EW_between, d_NS_between)
    pole_elevation = CorrectLengthAngle(network_length.length.values, el1, el2)
    network_length['elevation_angle'] = pole_elevation
    network_length['adj_length'] = network_length['length']/np.cos(np.radians(network_length.elevation_angle))
    
    from_poles, to_poles = LinesPoles(poleclasses, network_length)
    network_length['Pole_ID_From'] = from_poles
    network_length['Pole_ID_To'] = to_poles
    sub_nets, branches = NetworkLinesID(network_length, base)
    network_length['SubNetwork'] = sub_nets
    network_length['Branch'] = branches
    network_length = network_length[['index_x_from', 'index_y_from', 'index_x_to',
                                     'index_y_to', 'Type','Pole_ID_From', 'Pole_ID_To',
                                     'SubNetwork', 'Branch', 'length', 'elevation_angle',
                                     'adj_length', 'geometry']]
    
    connections = pd.DataFrame({'index_x':indexes_conn[:, 0], 'index_y':indexes_conn[:, 1]})
    geometry = [Point(p[0], p[1]) for p in T_associated_conns]
    connections = gpd.GeoDataFrame(connections, geometry=geometry)
    con_elv, con_utm_x, con_utm_y, con_gps = PoleElevation(gen_LON, gen_LAT, gen_site_indexes,
                                                           indexes_conn, d_EW_between, d_NS_between)
    connections['elevation'] = con_elv
    connections['UTM_X'] = con_utm_x
    connections['UTM_Y'] = con_utm_y
    connections['GPS_X'] = list(con_gps[:, 0])
    connections['GPS_Y'] = list(con_gps[:, 1])
    connections = connections[['index_x', 'index_y', 'UTM_X', 'UTM_Y', 'GPS_X',
                               'GPS_Y', 'elevation', 'geometry']]
    
    # Drop lines
    conn_fro_ids, drop_fro_ids = MatchConnectionsPoleID(
        poleclasses, droplines, connections, d_EW_between, d_NS_between)
    droplines['DropPoleID'] = drop_fro_ids
    dsub_net, dbranch_ = [], []
    for id_ in drop_fro_ids:
        rem = id_.replace(base, '')
        dsub_net.append(rem[0])
        dbranch_.append(rem[1])
    droplines['SubNetwork'] = dsub_net
    droplines['Branch'] = dbranch_
    droplines = droplines[['index_x_from', 'index_y_from', 'index_x_to', 'index_y_to',
                           'DropPoleID', 'SubNetwork', 'Branch', 'Linedrop', 'geometry']]
    
    return poleclasses, network_length, droplines, connections
#==============================================================================


#==============================================================================
# Evaluate characteristic Length
def CharacteristicLength(line_type, actual_length, char_length_factor):
    # if MV, return same
    if line_type == 'MV':
        return actual_length
    else:
        return actual_length*char_length_factor
#==============================================================================


#==============================================================================
# Evaluate current
def CurrentRating(line_type, t1_power, mv_3ph_voltage, connections, household_current,
                  other_power_factor, derate):
    if line_type == 'MV':
        return (t1_power/mv_3ph_voltage)/derate
    else:
        return (connections*household_current/other_power_factor)/derate
#==============================================================================


#==============================================================================
# Evaluate CB rating and the actual CB size
def CBSizing(current, breaker_rating, breaker_sizes):
    CBRating = current*breaker_rating
    idxs = np.where(breaker_sizes > CBRating)[0]
    CBSize =  breaker_sizes[idxs].min()
    return CBRating, CBSize
#==============================================================================


#=============================================================================
# Evaluate Voltage Drop 
def VoltageDrop(cable_type, char_length, current, voltage, max_drop_volt):
    RatedCurrentMatrix = np.array(
        [[27, 36, 45, 61, 78, 104, 125, 149, 190, 238, 272, 308, 350, 0, 0],
         [0,  0,  30, 50, 70, 0,   0,   0,   0,   0,   0,   0,   0, 0, 0],
         [0,  0,  30, 80, 100, 0,   0,   0,   0,   0,   0,   0,   0, 0, 0],
         [0,  0,  0,  0,  87,  111, 138, 168, 213, 258, 300, 339, 0, 0, 0],
         [0,  0,  0,  0,  1, 138, 168, 213, 258, 300, 339, 0, 0, 0, 0],
         [0,  0,  0,  0,  0,  96, 116, 138, 171, 205, 234, 263, 298, 344, 385],
         [0,  0,  0,  68, 91, 113, 0, 0, 0, 0, 0, 0, 0, 0, 0],
         [26, 34, 43, 58, 75, 95, 115, 140, 170, 205, 235, 265, 300, 345, 385],
         [0, 0, 0, 0, 0, 0, 0, 148, 0, 0, 0, 0, 0, 0, 0]])
    
    VoltageDropMatrix = np.array(
        [[0.018, 0.011, 0.0074, 0.0044, 0.0028, 0.0017, 0.0013, 0.00093, 0.00065, 
          0.00048, 0.00038, 0.00032, 0.00026, 0, 0],
         [0,  0,  0.0118, 0.0047, 0.0029, 0,   0,   0,   0,  0, 0, 0, 0, 0, 0],
         [0,  0,  0.0118, 0.0047, 0.0029, 0,   0,   0,   0,  0, 0, 0, 0, 0, 0],
         [0,  0,  0, 0, 0.0048, 0.0031, 0.0022, 0.0017, 0.0012, 0.0008, 0.0007,
          0.0006, 0, 0, 0],
         [0,  0, 0, 0, 0.0023, 0.0019, 0.0014, 0.0010, 0.0007, 0.0006, 0.0005,
          0, 0, 0, 0],
         [0,  0,  0,  0,  0,  0.0015, 0.0011, 0.0008, 0.0006, 0.0004, 0.0003, 
          0.0003, 0.0002, 0.0002, 0.0002],
         [0,  0,  0,  0.003793, 0.00239, 0.0012, 0, 0, 0, 0, 0, 0, 0, 0, 0],
         [0.0154, 0.0096, 0.0064, 0.0038, 0.0024, 0.0015, 0.0011, 0.0008, 0.0006,
          0.0004, 0.0004, 0.0003, 0.0003, 0.0002, 0.0002],
         [0, 0, 0, 0, 0, 0, 0, 0.0002, 0, 0, 0, 0, 0, 0, 0]])
    
    cable_types = ['GP', 'Airdac SNE', 'Airdac CNE', 'LV ABC (2-Core)', 'LV ABC (4-Core)', 
                   'PVC (1-Core) + BCEW', 'SWA (2-Core)', 'SWA (4-Core)', 'FOX ACSR']
    
    cable_sizes = [2.5, 4, 6, 10, 16, 25, 35, 50, 70]
    
    idx = cable_types.index(cable_type)
    vdrop = []
    for idy, cs in enumerate(cable_sizes):
        Irated = RatedCurrentMatrix[idx, idy]
        #print(Irated)
        if current > Irated or Irated == 0:
            vdrop.append("No Cable")
            #print(vdrop)
        else:
            vd = char_length*current*VoltageDropMatrix[idx, idy]/voltage
            if vd < max_drop_volt:
                vdrop.append("Pass")
                #print(vdrop)
            else:
                vdrop.append("High V-Drop")
                #print(vdrop)
    print(vd*230, (230 - vd*230), vd ,max_drop_volt)
    return vdrop
#==============================================================================



#==============================================================================
# Evaluate Cable Requirements
def CableCalculations(networklines, droplines, connections, char_length_factor,
                      t1_power, mv_3ph_voltage, household_voltage, household_current,
                      other_power_factor, derate, breaker_rating, max_drop_volt):
    
    # hard code breaker sizes:
    breaker_sizes = np.array([15, 20, 25 , 30, 35, 40, 45, 50, 60, 70, 80, 100,
                              110, 125, 150, 175, 200, 250, 300, 350, 400, 450,
                              500, 600, 700, 800, 1000, 1200, 1600, 2000, 2500,
                              3000, 4000, 5000, 6000])
    ntl = networklines.filter(items=['SubNetwork', 'Branch', 'adj_length'])
    ntl = ntl.rename(columns={'adj_length':'Length'})
    ntl = ntl.groupby(['SubNetwork', 'Branch']).sum()
   
    drp = droplines.filter(items=['SubNetwork', 'Branch', 'Linedrop'])
    drp = drp.rename(columns={'Linedrop':'Length'})
    drp.loc[len(drp)] = ['M', 1, 0]
    drp = drp.groupby(['SubNetwork', 'Branch']).sum()
    
    res_series = ntl + drp
    result = res_series.index.to_frame(index=False)
    result['Length'] = res_series.values
    
    con = droplines.filter(items=['SubNetwork', 'Branch'])
    con = con.value_counts()
    
    number_custs = []
    for index in zip(result.SubNetwork.values, result.Branch.values):
        if index in list(con.index):
            number_custs.append(con[index])
        else:
            number_custs.append(0)
    result['Connections'] = number_custs
    
    voltage, cable_type, ltype = [],[],[]
    for sub in result.SubNetwork.values:
        if sub == 'M':
            voltage.append(mv_3ph_voltage)
            cable_type.append('FOX ACSR')
            ltype.append('MV')
        else:
            voltage.append(household_voltage)
            cable_type.append('LV ABC (2-Core)')
            ltype.append('LV')
    result['CableType'] = cable_type
    result['Voltage'] = voltage
    result['LineType'] = ltype
    result = result[['SubNetwork', 'Branch', 'LineType', 'CableType', 'Voltage', 
                     'Connections', 'Length']]
    CL = []
    for ltyp, len_ in zip(result.LineType.values, result.Length.values):
        CL.append(CharacteristicLength(ltyp, len_, char_length_factor))
    result['CharacteristicLength'] = CL
    
    I = []
    for ltyp, cons in zip(result.LineType.values, result.Connections.values):
        I.append(CurrentRating(ltyp, t1_power, mv_3ph_voltage, cons, household_current,
                               other_power_factor, derate))
    result['Current'] = I
    
    cb_rating, actual_cb_size = [],[]
    for Ia in result.Current.values:
        cr, acbs = CBSizing(Ia, breaker_rating, breaker_sizes)
        cb_rating.append(cr)
        actual_cb_size.append(acbs)
    
    result['CB_Rating'] = cb_rating
    result['Actual_CB_Size'] = actual_cb_size
    
    cable_sizes = [2.5, 4, 6, 10, 16, 25, 35, 50, 70]
    volt_drop_df = pd.DataFrame(columns=['Cable Size ' + str(x) for x in cable_sizes])
    for i in result.index:
        df = result.loc[i]
        volt_drop_df.loc[i] = VoltageDrop(df.CableType, df.CharacteristicLength, df.Current, df.Voltage, max_drop_volt)
    res = pd.concat([result, volt_drop_df], axis=1)
    return res
#==============================================================================


#==============================================================================
# Put the results in an excel file 
def ConcessionDetails(dfpoles, dfnet, dfdropline, dfcosts, connections, voltagedropdf, 
                      concession, conc_id=None):

    path = os.path.join(os.getcwd(), concession) # directory to save in
    if os.path.isdir(path):
        pass
    else:
        os.mkdir(path)
    # Save these into excel
    wb = Workbook()
    wb.create_sheet(title="PoleClasses", index=0)
    wb.create_sheet(title="NetworkLength", index=1)
    wb.create_sheet(title="DropLines", index=2)
    wb.create_sheet(title="Connections", index=3)
    wb.create_sheet(title="NetworkCalculations", index=4)
    wb.create_sheet(title="NetworkLayout", index=5)
    wb.create_sheet(title='NetworkCost', index=6)
    ws = wb["PoleClasses"]
    for row in dataframe_to_rows(dfpoles.drop(columns=['geometry']), header=True):
        ws.append(row)
    ws = wb["NetworkLength"]
    for row in dataframe_to_rows(dfnet.drop(columns=['geometry']), header=True):
        ws.append(row)
    ws = wb['DropLines']
    for row in dataframe_to_rows(dfdropline.drop(columns=['geometry']), header=True):
        ws.append(row)
    ws = wb['Connections']
    for row in dataframe_to_rows(connections.drop(columns=['geometry']), header=True):
        ws.append(row)
    ws = wb['NetworkCalculations']
    for row in dataframe_to_rows(voltagedropdf, header=True):
        ws.append(row)
    # Evaluate date and time on which the simulation is done
    simdate = dt.datetime.today() # simulation date

    # Define a function that ensures double digits for day of the month
    add0 = lambda x: '0'+str(x) if x < 10 else str(x)
    
    filename = str(simdate.year) + add0(simdate.month) + add0(simdate.day) + \
        '_' + add0(simdate.hour) + add0(simdate.minute) + '_' + concession.upper()[0:3] + \
        '_uGrid'
    ws = wb["NetworkLayout"]
    fig, ax = plt.subplots(dpi=250)
    dfpoles[dfpoles.Type == "MV"].plot(column='Type', s=8, ax=ax, legend=False)
    dfpoles[dfpoles.Type == "LV"].plot(column='Type', s=4, ax=ax, legend=False)
    connections.plot(markersize=2, color='brown', ax=ax, legend=False)
    dfdropline.plot(lw=0.5, ax=ax, color='red', legend=False)
    dfnet[dfnet.Type != "MV"].plot(column="Type", lw=0.7,ax=ax,legend=False)
    dfnet[dfnet.Type == "MV"].plot(lw=1,ax=ax, color='blue',legend=False)
    plt.savefig(concession+'.png')
    ws.add_image(img(concession+'.png'), 'B4')
    
    ws = wb['NetworkCost']
    for row in dataframe_to_rows(dfcosts, index=True, header=True):
        ws.append(row)
    free_cell_loc = len(dfcosts)+3
    ws['D'+str(free_cell_loc)] = 'Total'
    ws['E'+str(free_cell_loc)] = dfcosts['Line Total (USD)'].values.sum()
    if conc_id == None:
        wb.save(path+'/'+ filename + ".xlsx")
    else:
        wb.save(path+'/'+ filename + str(conc_id)+ ".xlsx")
    
    # Save shapefiles
    shapepath = os.path.join(path, 'GIS_Files')
    if os.path.isdir(shapepath):
        pass
    else:
        os.mkdir(shapepath)
    dfpoles.to_file(shapepath + '/poles.shp')
    dfnet.to_file(shapepath + '/network.shp')
    dfdropline.to_file(shapepath + '/droplines.shp')
    connections.to_file(shapepath + '/customers.shp')
    
#==============================================================================

#==============================================================================
def AddSpur(concession, filename, gen_LON, gen_LAT, gen_indexes, indexes, type_,
            d_EW_between, d_NS_between, costs, conc_id=None):
    path = os.path.join(os.getcwd(), concession) # directory
    try:
        poleclasses = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='PoleClasses', skiprows=[1])
        networklength = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='NetworkLength', skiprows=[1])
        droplines = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='DropLines', skiprows=[1])
        conns = pd.read_excel(path+'/'+filename+".xlsx", sheet_name='Connections', skiprows=[1])
    except:
        poleclasses = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='PoleClasses', skiprows=[1])
        networklength = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='NetworkLength', skiprows=[1])
        droplines = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='DropLines', skiprows=[1])
        conns = pd.read_excel(path+'/'+filename+str(conc_id)+".xlsx", sheet_name='Connections', skiprows=[1])
    poleclasses = poleclasses.drop(columns= ['Unnamed: 0'])
    pclasses = poleclasses[poleclasses.Type == type_.upper()]
    
    pole_indexes = pclasses.filter(items=['index_x', 'index_y']).to_numpy()
    pole_ids = pclasses.ID.values
    d_bw_poleidxs_newidxs = DistanceBWindexes(pole_indexes, indexes, d_EW_between, d_NS_between)
    min_dstanc = d_bw_poleidxs_newidxs.min()
    x_index, y_index = np.where(d_bw_poleidxs_newidxs == min_dstanc)
    base_ID = pole_ids[x_index[0]]
    conn_pole = pole_indexes[x_index[0], :]
    source = np.array([conn_pole[0]*d_EW_between, conn_pole[1]*d_NS_between])
    X = indexes*np.array([d_EW_between, d_NS_between])
    dists = np.sqrt(np.sum(np.power(X - source, 2), axis=1))
    dists_copy = np.copy(dists)
    dists_copy.sort()
    C = np.zeros((len(indexes), 2))
    ordered_index = [np.where(dists == val)[0][0] for val in dists_copy]
    C[:, 0:2] = indexes[ordered_index]
    indexes_IDs = []
    for i in range(len(C)):
        indexes_IDs.append(base_ID + 'S'+str(i+1))
    nw_indxs = np.vstack([C, conn_pole])
    nw_classes = PoleAngleClass(nw_indxs, d_EW_between, d_NS_between, type_)
    nw_classes = nw_classes[:-1]
    el, utmx, utmy, gps = PoleElevation(gen_LON, gen_LAT, gen_indexes, C, d_EW_between, d_NS_between)
    nw_classes['distance_from_source'] = dists_copy
    nw_classes['ID'] = indexes_IDs
    nw_classes['elevation'] = el
    nw_classes['UTM_X'] = utmx
    nw_classes['UTM_Y'] = utmy
    nw_classes['Type'] = type_
    nw_classes['GPS_X'] = gps[:, 0]
    nw_classes['GPS_Y'] = gps[:, 1]
    new_poleclasses = pd.concat([poleclasses, nw_classes])
    
    GDF = gpd.GeoDataFrame(new_poleclasses, 
                           geometry=gpd.points_from_xy(new_poleclasses.index_x*d_EW_between,
                                                       new_poleclasses.index_y*d_NS_between))
    
    from_idxs_net = networklength.filter(items=['index_x_from', 'index_y_from']).to_numpy()
    to_idxs_net = networklength.filter(items=['index_x_to', 'index_y_to']).to_numpy()
    from_coords = from_idxs_net*np.array([d_EW_between, d_NS_between])
    to_coords = to_idxs_net*np.array([d_EW_between, d_NS_between])
    Lines = [LineString((from_coords[i, :], to_coords[i, :])) for i in range(len(to_coords))]
    if type_ == 'MV':
        net_len = NetworkLength(nw_indxs, d_EW_between, d_NS_between)
        net_len['Type'] = 'MV'
    else:
        ntlength = networklength[networklength.Type != 'MV']
        id_ = None
        for typ in ntlength.Type.unique():
            nt = ntlength[ntlength.Type == typ]
            
            from_ = np.zeros((len(nt), 2))
            from_[:, 0] = nt.index_x_from.values
            from_[:, 1] = nt.index_y_from.values
            
            to_ = np.zeros((len(nt), 2))
            to_[:, 0] = nt.index_x_to.values
            to_[:, 1] = nt.index_y_to.values
            
            from_minus = np.abs(from_ - conn_pole)
            to_minus = np.abs(to_ - conn_pole)
            
            if from_minus.min(axis=0)[0] == 0 and from_minus.min(axis=0)[1] == 0:
                id_ = typ
                break
            if to_minus.min(axis=0)[0] == 0 and to_minus.min(axis=0)[1] == 0:
                id_ = typ
                break
        net_len = NetworkLength(nw_indxs, d_EW_between, d_NS_between)
        net_len['Type'] = id_
        
     # Put nework lines into a dataframe and geodataframe
    from_indexes = np.zeros((len(net_len), 2))
    from_indexes[:, 0] = net_len.index_x_from.values
    from_indexes[:, 1] = net_len.index_y_from.values
    
    to_indexes = np.zeros((len(net_len), 2))
    to_indexes[:, 0] = net_len.index_x_to.values
    to_indexes[:, 1] = net_len.index_y_to.values
    el1, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_indexes, from_indexes, d_EW_between, d_NS_between)
    el2, utm_x, utm_y, gps = PoleElevation(gen_LON, gen_LAT, gen_indexes, to_indexes, d_EW_between, d_NS_between)
    pole_elevation = CorrectLengthAngle(net_len.length.values, el1, el2)
    net_len['elevation_angle'] = pole_elevation
    net_len['adj_length'] = net_len['length']/np.cos(np.radians(net_len.elevation_angle))
    
    networklength = gpd.GeoDataFrame(networklength, geometry=Lines)
    net_len = pd.concat([networklength, net_len])
    
    networkcost = NetworkCost(costs, poleclasses, networklines, droplines)
    
    drop_fro = droplines.filter(items=['index_x_from', 'index_y_from']).to_numpy()
    drop_to = droplines.filter(items=['index_x_to', 'index_y_to']).to_numpy()
    
    drop_fro = drop_fro*np.array([d_EW_between, d_NS_between])
    drop_to = drop_to*np.array([d_EW_between, d_NS_between])
    
    lines = [LineString((drop_fro[i, :], drop_to[i, :])) for i in range(len(drop_to))]
    droplines = gpd.GeoDataFrame(droplines, geometry=lines)
    
    indexes_conn = conns.to_numpy()
    T_associated_conns = indexes_conn*np.array([d_EW_between, d_NS_between])
    connections = pd.DataFrame({'index_x':indexes_conn[:, 0], 'index_y':indexes_conn[:, 1]})
    geometry = [Point(p[0], p[1]) for p in T_associated_conns]
    connections = gpd.GeoDataFrame(connections, geometry=geometry)
    
    ConcessionDetails(GDF, net_len, droplines, networkcost, connections, concession, conc_id)
#==============================================================================

def gpsRecalibration(gen_lat, un_lat, gen_long, un_long, orig_long = -1, orig_lat = -1):
    #RECALIBRATE THE GPS COORDINATES FOR THE UGRIDNET OUTPUT
    #Constraints: 1) Ideally Requires the correlation equation be changed respectively for every site run 
    #             2) If need be, the equation can be created using excel spreadsheet

    #TODO: ADD A MORE ROBUST FLATTENING PROJECTION FIX FOR THE COORDINATES
    
    ref_lat, ref_long = gen_lat,gen_long
    
    #lat_offset = orig_lat - un_lat
    #long_offset  =  orig_long - un_long

    if(orig_long != -1 and orig_lat != -1): 
        #used if the original values/coordinates are known
        average_lat = 0.5*(orig_lat + un_lat)
        average_lat_offset = average_lat - ref_lat
        average_long = 0.5*(orig_long + un_long)
        average_long_offset = average_long - ref_long
    else:
        
        average_lat_offset = un_lat - ref_lat
        average_long_offset = un_long - ref_long

    #Equations to correct/calibrate uGridNet coordinates outputs
    corr_lat = (0.005065190225781*abs(average_lat_offset) - 0.00000283805355199712) + un_lat # correction of latitude
    corr_long = ((0.003791523859891*average_long_offset) + 0.00000284320752421793) + un_long #correction of longitude
    

    return corr_lat, corr_long



#==============================================================================
# Run the Network 
def SimulateNetwork(site_properties, conc_ID=None, min_trans=1):
    # Collect village details
    indexes_conn = site_properties['indexes_conn']
    indexes_excl = site_properties['indexes_excl']
    max_y = site_properties['height']
    max_x = site_properties['width']
    d_EW_between = site_properties['d_EW_between']
    d_NS_between = site_properties['d_NS_between']
    Long_exc_min = site_properties['Long_exc_min']
    Long_exc_max = site_properties['Long_exc_max']
    Lat_exc_min = site_properties['Lat_exc_min']
    Lat_exc_max = site_properties['Lat_exc_max']
    load_per_conn = site_properties['load_per_conn']
    kW_max = site_properties['peak_demand']
    reformatScaler = site_properties['reformatScaler']
    exclusionBuffer = site_properties['exclusionBuffer']
    MaxDistancePoleConn = site_properties["MaxDistancePoleConn"]
    MaxDistancePoleLV = site_properties["MaxDistancePoleLV"]
    MaxDistancePoleMV = site_properties["MaxDistancePoleMV"]
    range_limit = site_properties["range_limit"]
    lat_Generation = site_properties["lat_Generation"]
    long_Generation = site_properties["long_Generation"]
    derate = site_properties["derate"]
    tl_power = site_properties["tl_power"]
    other_power_factor = site_properties["other_power_factor"]
    mv_3ph_voltage = site_properties["mv_3ph_voltage"]
    lv_3ph_voltage = site_properties["lv_3ph_voltage"]
    household_voltage = site_properties["household_voltage"]
    max_drop_volt = site_properties["max_drop_volt"]
    breaker_rating_factor = site_properties["breaker_rating_factor"]
    char_length_factor = site_properties["characteristic_length_factor"]
    costs = site_properties["costs"]
    
    # Evaluate minimum number of transformers
    LV_kW = (230*130)/1000 # kW - LV line 
    min_num_trans = max(int(kW_max/LV_kW), min_trans, len(indexes_conn)//50)
    if len(indexes_conn) < 20:
        max_num_trans = min_num_trans + 1
    elif len(indexes_conn) in range(20,60):
        max_num_trans = min_num_trans + 3
    elif len(indexes_conn) in range(60,100):
        max_num_trans = min_num_trans + 5
    elif len(indexes_conn) in range(100,140):
        max_num_trans = min_num_trans + 8
    else:
        max_num_trans = min_num_trans + 15
    
    household_current = 1.2*(kW_max*1000/(len(indexes_conn)*230))
    if household_current < 1.712262:
        household_current = 2
    print("Household current is {}".format(household_current))
    
    # Set number of repeats 
    num_repeats = 5
    BestCost = 999999999999999 # dummy cost value
    BestPoleClasses = None
    BestNetworkLines = None
    BestDropLines = None
    BestVoltageDrop = None
    for i in range(num_repeats):
        # Evaluate the network for all transformers from mimimun to maximum
        for num_trans in range(min_num_trans, max_num_trans):
            print("Iteration number {} of {}, with {} transformers".format(i+1, num_repeats, num_trans))
            # Evaluate the customer clusters
            min_custs = 3 # No transformer can have less than 3 customers
            if min_custs*num_trans > len(indexes_conn): 
                print("Number of customers, {} too low to justify {} transformers".format(len(indexes_conn), num_trans))
                break
            max_custs = min(len(indexes_conn)//num_trans + 25, len(indexes_conn))
            conn_clusters, centroids = ClusterTransformerCusts(indexes_conn, d_EW_between, 
                                                    d_NS_between, num_trans, min_custs,
                                                    max_custs)
            
            # Place LV Poles 
            LV_Pole_indexes = LVPolesPlacement(conn_clusters, d_EW_between, d_NS_between, 
                                                indexes_excl, range_limit, max_x, max_y,
                                                MaxDistancePoleConn)
            
            # Evaluate extended LV connection Poles
            LV_Pole_extended = IntermediateConnectionLVPoles(conn_clusters, LV_Pole_indexes,
                                                              d_EW_between, d_NS_between,
                                                              indexes_excl, range_limit,
                                                              max_x, max_y, MaxDistancePoleLV)
            
            Trans_Locs = TransformerLocations(LV_Pole_extended, d_EW_between, d_NS_between, 
                                              indexes_excl, range_limit, max_x, max_y)
            
            # Place the transformer location into LV pole indexes
            for idx, tloc in enumerate(Trans_Locs):
                LV_Pole_extended[idx] = np.vstack([LV_Pole_extended[idx], tloc])
                
            # Evaluate intermediate LV Poles 
            all_LV_Poles = []
            for g in LV_Pole_extended:
                gdf = pd.DataFrame({'x':g[:, 0], 'y':g[:, 1]})
                gdf = gdf.drop_duplicates(subset=['x', 'y'])
                g = gdf.to_numpy()
                OnOff = MSTConnectivityMatrix(g, d_EW_between, d_NS_between)
                distances = DistanceBWindexes(g, g, d_EW_between, d_NS_between)
                g_new = IntermediatePoles(g, OnOff, distances, MaxDistancePoleLV, 
                                          indexes_excl, range_limit, max_y, max_x)
                all_LV_Poles.append(g_new)
                
            #Calculate HV Wiring Layout
            #Find POI of generation
            ccc = POI_Pole(lat_Generation,long_Generation,Long_exc_min,Lat_exc_min,
                           d_EW_between,d_NS_between,Trans_Locs, MaxDistancePoleMV)
            closest_pole, indexes_gen, connection_point = ccc
            
            # add POI_MV to wiring for MV
            indexes_Poles_MV_wPOI = np.vstack([Trans_Locs, indexes_gen])
            
            # Evaluate OnOffMV poles
            OnOff_MV = MSTConnectivityMatrix(indexes_Poles_MV_wPOI, d_EW_between, d_NS_between)
            
            # Evaluate distances between MV Poles
            DistancesBWPoles_MV = distance_matrix(indexes_Poles_MV_wPOI*np.array([d_EW_between, d_NS_between]),
                                                  indexes_Poles_MV_wPOI*np.array([d_EW_between, d_NS_between]))
            
            # Evaluate OnOffLV poles
            OnOff_LV = []
            for LV_Poles in all_LV_Poles:
                OnOff = MSTConnectivityMatrix(LV_Poles, d_EW_between, d_NS_between)
                OnOff_LV.append(OnOff)
            
            # Classify Network Poles
            CL = ClassifyNetworkPoles(concession, lat_Generation, long_Generation,
                                      indexes_gen, indexes_Poles_MV_wPOI, all_LV_Poles,
                                      OnOff_MV, DistancesBWPoles_MV, indexes_excl,
                                      MaxDistancePoleMV, range_limit, max_x, max_y, 
                                      d_EW_between, d_NS_between, indexes_conn, conc_ID)
            
            poleclasses, networklines, droplines, connections = CL
            voltage_drop_df = CableCalculations(networklines, droplines, connections,
                                                char_length_factor, tl_power, mv_3ph_voltage, 
                                                household_voltage, household_current,
                                                other_power_factor, derate,
                                                breaker_rating_factor, max_drop_volt)
            network_fail = []
            for j in voltage_drop_df.index:
                vals = voltage_drop_df.loc[j].values
                if 'Pass' in vals:
                    pass
                else:
                    network_fail.append('Fail')
            networkcost = NetworkCost(costs, poleclasses, networklines, droplines)
            cost_total = networkcost['Line Total (USD)'].values.sum()
            print('Cost: $',cost_total)
            if cost_total < BestCost and not network_fail:
                print('Network Passed')
                BestCost = cost_total.copy()
                BestNetworkCost = networkcost.copy()
                BestDropLines = droplines.copy()
                BestNetworkLines = networklines.copy()
                BestPoleClasses = poleclasses.copy() 
                BestVoltageDrop = voltage_drop_df.copy()
    # Put this into excel file
    if BestCost < 999999999999999:
        print("Save Best Solution")
        ConcessionDetails(BestPoleClasses, BestNetworkLines, BestDropLines, BestNetworkCost,
                      connections, BestVoltageDrop, concession, conc_ID)
    else:
        print("Could not find a solution, trying again!")
        return SimulateNetwork(site_properties, conc_ID, min_trans + 5)
    return BestPoleClasses, BestNetworkLines, BestDropLines, connections, BestVoltageDrop, BestNetworkCost
#==============================================================================




if __name__=='__main__':
    
    if os.path.exists('indexes_conn_reformatted_1.csv'):
        os.remove('indexes_conn_reformatted_1.csv')
    
    if os.path.exists('indexes_reformatted_1_bufferzone_2.csv'):
        os.remove('indexes_reformatted_1_bufferzone_2.csv')
    
    if os.path.exists('index_maxes_1.csv'):
        os.remove('index_maxes_1.csv')
        
    if os.path.exists('d_between_1.csv'):
        os.remove('d_between_1.csv')
        
    site_properties = CollectVillageData(site_name)
    # for k, v in site_properties.items():
    #     exec(k + '= v')
    
    print("Done collecting village info, run optimizer")
    poleclasses, networklines, droplines, connections, vd, nc = SimulateNetwork(site_properties, conc_ID=concession_id)
    
    fig, ax = plt.subplots(dpi=150)
    poleclasses[poleclasses.Type == 'MV'].plot(markersize=1, color='red', ax=ax)
    poleclasses[poleclasses.Type == 'LV'].plot(markersize=0.5, color='blue', ax=ax)
    networklines[networklines.Type =='MV'].plot(linewidth=0.5, color='red', ls='--', ax=ax)
    networklines[networklines.Type != 'MV'].plot(linewidth=0.5, color='blue', ls='--', ax=ax)
    connections.plot(markersize=0.5, color='green', ax=ax)
    droplines.plot(linewidth=1, ls=':', color='green', ax=ax)
    plt.show()
    
    # cc, c = ClusterTransformerCusts(indexes_conn, d_EW_between, d_NS_between, 10, 3, 60)
    # fig, ax = plt.subplots(dpi=200)
    # for cc_ in cc:
    #     print(len(cc_))
    #     ax.scatter(cc_[:, 0], cc_[:, 1], s=5)
    
